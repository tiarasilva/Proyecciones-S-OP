from constants import *
from tkinter import *
from tkinter import ttk, messagebox
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl import load_workbook

def messageBox(dict_lead_time, selected_tipo_venta):
  app = Tk()
  app.configure(bg='#ffffff')

  # Info Image
  information_image = Image.open("Img/Notice.png")
  information_image = information_image.resize((50, 50))
  information_image = ImageTk.PhotoImage(information_image)
  logo_label = tk.Label(image = information_image, borderwidth=0, bg = 'white')
  logo_label.image = information_image
  logo_label.grid(column = 0, row = 1, padx = 10, pady = 10)

  # Instrucciones
  instructions = tk.Label(app, text = "Stock detenido", bg="white", justify=LEFT)
  instructions.configure(font=("bold", 16))
  instructions.grid(column = 1, row = 1)

  # Description
  messageboxText = "Los siguientes Stocks aún no han sido liberados, a pesar de que hayan sobrepasado su tiempo de lead time."
  text = Text(app, bg = '#ffffff', bd = 0, borderwidth=0, highlightthickness=0, height=3, width=65, pady = 2)
  text.insert(INSERT, messageboxText)
  text.config(state=DISABLED, font=("Calibrí", 13))
  text.grid(column = 0, row = 2, columnspan = 12, padx=20)

  # Table
  tree = ttk.Treeview(app, columns=(1, 2, 3, 4, 5), show="headings", height="7", padding=(10,0,0,10))
  tree.grid(column = 0, row = 3, columnspan = 12)

  scrollbar = ttk.Scrollbar(app, orient=tk.VERTICAL, command=tree.yview)
  tree.config(yscroll = scrollbar.set)
  scrollbar.grid(column=12, row=3, sticky='ns', padx=10)

  tree.column(1, anchor=CENTER, stretch=NO, width=70)
  tree.heading(1, text = 'Fila')
  tree.column(2, anchor=CENTER, stretch=NO, width=200)
  tree.heading(2, text = 'Llave')
  tree.column(3, anchor=CENTER, stretch=NO, width=120)
  tree.heading(3, text = 'KG')
  tree.column(4, anchor=CENTER, stretch=NO, width=100)
  tree.heading(4, text = 'Días oficina')
  tree.column(5, anchor=CENTER, stretch=NO, width=100)
  tree.heading(5, text = 'Lead time Destino')

  # Data
  wb_dias_stock = load_workbook(filename_dias, read_only=True, data_only=True)
  ws_dias_stock = wb_dias_stock.active
  sector = ''
  oficina = ''

  for i, row in enumerate(ws_dias_stock.iter_rows(5, ws_dias_stock.max_row, values_only=True), 4):
    if row[2] is None:
      break
    
    if row[0] is not None:
      sector = row[0]
    
    if row[1] is not None:
      oficina = row[1]
    
    material = row[2]
    descripcion = row[3]

    stock_oficina_no_lib = row[16] or 0
    dias_oficina_centro_no_lib = row[17] or 0
    dias_oficina_no_lib = row[11] or 0

    lead_time = dict_lead_time['optimista'][selected_tipo_venta][oficina.lower()]

    if dias_oficina_centro_no_lib >= lead_time['Destino']:
      tree.insert('', 'end', values = (i, f'{oficina.lower()}{material}', stock_oficina_no_lib, round(dias_oficina_centro_no_lib, 1), round(lead_time['Destino'], 1)))
  
  wb_dias_stock.close() 

  # Exit button
  boton = Button(app, text = "Salir", command = app.destroy, width=8, highlightbackground='#ffffff')
  boton.grid(column = 10, row = 6, pady = 20)


  app.title('Stock detenido')
  app.eval('tk::PlaceWindow . center')
  app.mainloop()