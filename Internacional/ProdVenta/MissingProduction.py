from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from constants import *

def create_missing_production(ws, dict_leftover_country):
  ws.append({
    1: 'Fecha',                     # A
    2: 'Sector',                    # B
    3: 'Oficina',                   # C
    4: 'Material',                  # D
    5: 'Descripción',               # E
    6: 'Nivel 2',                   # F
    7: 'Llave',                     # G
    8: 'Plan',                      # H
    9: 'Prod. actual',              # I
    10: 'Resta',                    # J
    11: 'Días productivos',         # K
    13: 'Lead time',                # L
    14: 'Porcentaje productivo',    # M
    15: 'Producción total posible', # N
  })

  # STYLES
  thin = Side(border_style="thin", color=white)
  for i in range(1, ws.max_column + 1):
    ws[f'{get_column_letter(i)}1'].font = Font(bold=True, color=white)
    ws[f'{get_column_letter(i)}1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'{get_column_letter(i)}1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{get_column_letter(i)}1'].fill = PatternFill("solid", fgColor=lightBlue)

  # canal_distribucion = 'Venta Directa'
  # holidays_country = dict_holidays[month_1.year]['chile']

  # if oficina is not None:
  #   if oficina.lower() in dict_lead_time['optimista']['Venta Local'].keys():
  #     canal_distribucion = 'Venta Local'
  #     holidays_country = dict_holidays[month_1.year][oficina.lower()]

  # if oficina.lower() in dict_lead_time['optimista'][canal_distribucion]:
  #   lead_time_opt = dict_lead_time['optimista'][canal_distribucion][oficina.lower()]
  #   lead_time_pes = dict_lead_time['pesimista'][canal_distribucion][oficina.lower()]

  # LT_pes = lead_time_pes['Puerto']
  #   LT_opt = lead_time_opt['Puerto']
  #   today = datetime.now()
  #   last_day_month = calendar.monthrange(today.year, today.month)[1]
  #   dict_leftover_country = {}
  
  # # Leftover days
  # leftover_days = 0
  # if oficina in dict_leftover_country:
  #   leftover_days = dict_leftover_country[oficina.lower()]
  # else:
  #   for day in range(today.day + 1, last_day_month + 1):
  #     date_day = date(today.year, today.month, day)
  #     if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
  #       leftover_days += 1
  #   dict_leftover_country[oficina.lower()] = leftover_days
  
  # # Porcentaje producción
  # pct_prod_pes = max(leftover_days - LT_pes, 0) / leftover_days
  # pct_prod_opt = max(leftover_days - LT_opt, 0) / leftover_days