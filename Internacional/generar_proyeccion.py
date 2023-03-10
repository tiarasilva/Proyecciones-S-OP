from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, BUILTIN_FORMATS
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from constants import *
from ETA.ETA import create_ETA
from MessageBox.MessageBox import messageBox
from PuertoChile.PuertoChile import create_puerto_chile
# from ProdVenta.MissingProduction import create_missing_production
from Stock.stock import stock
from styles import run_styles

import time
import calendar
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import holidays
from dateutil import rrule

import sys, os
from os import path

start_time = time.time()

# ----- 0. PATH
if getattr(sys, 'frozen', False):
  print('\nRunning in a PyInstaller bundle\n')
  bundle_dir = sys._MEIPASS
  filename_parametros = path.abspath(path.join(path.dirname(__file__), filename_parametros))
  filename_venta = path.abspath(path.join(path.dirname(__file__), filename_venta))
  filename_asignaciones = path.abspath(path.join(path.dirname(__file__), filename_asignaciones))
  filename_chile = path.abspath(path.join(path.dirname(__file__), filename_chile))
  filename_dias = path.abspath(path.join(path.dirname(__file__), filename_dias))
  filename_logistica = path.abspath(path.join(path.dirname(__file__), filename_logistica))
  filename_pedidos_confirmados = path.abspath(path.join(path.dirname(__file__), filename_pedidos_confirmados))
  filename = path.abspath(path.join(path.dirname(sys.executable), filename))
  path_img = path.abspath(path.join(path.dirname(__file__), path_img))

else:
  print('\nRunning in a Python environment\n')
  bundle_dir = os.path.dirname(os.path.abspath(__file__))

# # ----- 1. Abrimos el excel de los parametros
wb_parametros = load_workbook(filename_parametros, data_only = True, read_only = True)
ws_parametros_time = wb_parametros['Lead time']
ws_parametros_venta = wb_parametros['Venta']
ws_porcentaje = wb_parametros['Asignación productivo']

# DIAS PARA VENDER
ws_parametros_dias_venta = wb_parametros['Cierre venta']
dict_cierre_venta = {}
ws_dias_max = ws_parametros_dias_venta.max_row

for row in ws_parametros_dias_venta.iter_rows(2, ws_dias_max, values_only = True):
  if row[1] is None:
    break
  oficina = row[1]
  dias_cierre = row[2]
  dict_cierre_venta[oficina.lower()] = dias_cierre

# FECHAS
selected_year = ws_parametros_venta['B1'].value
selected_month = ws_parametros_venta['B2'].value
selected_week = ws_parametros_venta['B3'].value
number_selected_month = month_number[selected_month.lower()]
today = datetime.now().date()

last_day_selected_month = calendar.monthrange(selected_year, number_selected_month)[1]
date_selected_month = date(int(selected_year), number_selected_month, last_day_selected_month)
last_day_month = calendar.monthrange(today.year, today.month)[1]


# Nombre fechas
month_1 = date_selected_month
month_2 = date_selected_month + relativedelta(months=1)
month_3 = date_selected_month + relativedelta(months=2)

name_month_1 = month_translate_EN_CL[month_1.strftime('%B').lower()]
name_month_2 = month_translate_EN_CL[month_2.strftime('%B').lower()]
name_month_3 = month_translate_EN_CL[month_3.strftime('%B').lower()]

selected_month1_year = f'{month_1.strftime("%m")}.{month_1.year}'
selected_month2_year = f'{month_2.strftime("%m")}.{month_2.year}'
selected_month3_year = f'{month_3.strftime("%m")}.{month_3.year}'

# HOLIDAYS
last_year = month_1.year - 1
this_year = month_1.year
next_year = month_1.year + 1

dict_holidays = {
  last_year: {
    'agro america': holidays.US(years=last_year),         # USA
    'agro europa': holidays.IT(years=last_year),          # ITALIA
    'agro mexico': holidays.MX(years=last_year),          # Mexico
    'agrosuper shanghai': holidays.CN(years=last_year),   # China
    'andes asia': holidays.KR(years=last_year),           # Corea del sur
    'chile': holidays.CL(years=last_year)
  },
  today.year: {
    'agro america': holidays.US(years=this_year),         # USA
    'agro europa': holidays.IT(years=this_year),          # ITALIA
    'agro mexico': holidays.MX(years=this_year),          # Mexico
    'agrosuper shanghai': holidays.CN(years=this_year),   # China
    'andes asia': holidays.KR(years=this_year),           # Corea del sur
    'chile': holidays.CL(years=this_year)
  },
  next_year: {
    'agro america': holidays.US(years=next_year),         # USA
    'agro europa': holidays.IT(years=next_year),          # ITALIA
    'agro mexico': holidays.MX(years=next_year),          # Mexico
    'agrosuper shanghai': holidays.CN(years=next_year),   # China
    'andes asia': holidays.KR(years=next_year),           # Corea del sur
    'chile': holidays.CL(years=next_year)
  }
}

# TIPO DE VENTA y LEAD TIME
venta_iteracion = 'Venta Directa'
escenario_iteracion = 'optimista'
dict_lead_time = {
  'optimista': {'Venta Directa': {}, 'Venta Local': {}},
  'pesimista': {'Venta Directa': {}, 'Venta Local': {}},
}

for row in ws_parametros_time.iter_rows(2, ws_parametros_time.max_row, values_only = True):
  if 'Venta Local' == row[0]:
    venta_iteracion = 'Venta Local'
  
  if 'Venta Directa' == row[0]:
    venta_iteracion = 'Venta Directa'
  
  if 'PESIMISTA' == row[0]:
    escenario_iteracion = 'pesimista'

  oficina = row[1]
  planta = row[2]
  puerto = row[3]
  agua = row[4]
  destino = row[5]
  almacen = row[6]

  if oficina is not None:
    if oficina.lower() != 'oficina':
      if 'Venta Local' == venta_iteracion:
        dict_lead_time[escenario_iteracion][venta_iteracion][oficina.lower()] = { 'Planta': planta, 'Puerto': puerto, 'Agua': agua, 'Destino': destino, 'Almacen': almacen }
      else:
        dict_lead_time[escenario_iteracion][venta_iteracion][oficina.lower()] = { 'Planta': planta, 'Puerto': puerto }

# PORCENTAJE DE PRODUCCIÓN
dict_porcentaje_produccion = {}

for row in ws_porcentaje.iter_rows(2, ws_porcentaje.max_row, values_only=True):
  oficina = row[1].lower()
  produccion = row[2]
  dict_porcentaje_produccion[oficina] = produccion

wb_parametros.close()

# LEFTOVER POR OFICINAS
dict_leftover_country = {
  'chile': 0,
  'agro america' : 0,
  'agro europa' : 0,
  'agro mexico' : 0,
  'agrosuper shanghai' : 0,
  'andes asia' : 0,
}


# LEFTOVERDAYS BY SELECTED MONTH AND YEAR 
if today > month_1:
  print(f'La fecha seleccionada {name_month_1} {month_1.year} ya pasó el mes actual')

else:
  print(f'La fecha seleccionada es {name_month_1} {month_1.year}')
  for oficina in dict_leftover_country.keys():
    leftover_days = 0
    holidays_country = dict_holidays[month_1.year][oficina.lower()]

    for day in rrule.rrule(rrule.DAILY, dtstart=today, until=month_1):
      day = day.date()
      if day not in holidays_country and day.strftime('%A') != 'Sunday':
        leftover_days += 1
    dict_leftover_country[oficina.lower()] = leftover_days

# PRODUCTIVE DAYS IN CHILE until selected date
productive_days = 0
holidays_country = dict_holidays[month_1.year]['chile']
start_date = date(today.year, today.month, 1)

for day in rrule.rrule(rrule.DAILY, dtstart=start_date, until=month_1):
  if day not in holidays_country and day.strftime('%A') != 'Sunday':
    productive_days += 1

print(f'Días productivos totales desde el inicio de mes {start_date} hasta {month_1} son {productive_days}')

# ----- 2. Creamos el excel de resultados VENTA LOCAL y VENTA DIRECTA
wb = Workbook()
ws = wb.active
ws.title = sheet_name

ws.append({
  12: f'Proyección {name_month_1} {month_1.year}',
  24: f'Proyección {name_month_2} {month_2.year}',
  30: f'Proyección {name_month_3} {month_3.year}'
})

ws.append({
  1: 'Sector',                            # A
  2: 'Canal de Distribución',             # B
  3: 'Llave',                             # C
  4: 'Oficina',                           # D
  5: 'Material',                          # E
  6: 'Descripción',                       # F
  7: 'Nivel 2',                           # G
  8: 'Venta Actual',                      # H
  9: 'Plan total KG',                     # I
  10: 'RV Final Producción [KG]',         # J
  11: 'RV Final Venta [KG]',              # K

  12: 'Plan - Prod. actual',              # L
  13: 'ETA Pesimista',                    # M
  14: 'Puerto Chile Pes.',                # N
  15: 'Puerto Oficina Pes.',              # O
  16: 'Almacen oficina Pes.',             # P
  17: 'PROY. Pesimista',                  # Q

  18: 'Plan - Prod. actual',              # R
  19: 'ETA Optimista',                    # S
  20: 'Puerto Chile Opt.',                # T
  21: 'Puerto Oficina Opt.',              # U
  22: 'Almacen oficina Opt.',             # V
  23: 'PROY. Optimista',                  # W

  24: 'RV Produccion',                    # X
  25: 'RV Venta',                         # Y
  26: 'ETA Pesimista N+1',                # Z
  27: f'Proy. Pesimista {name_month_2}',  # AA
  28: 'ETA Optimista N+1',                # AB
  29: f'Proy. Optimista {name_month_2}',  # AC

  30: 'Asignación de venta',              # AD
  31: 'ETA Pesimista N+2',                # AE
  32: f'Proy. Pesimista {name_month_3}',  # AF
  
  33: 'ETA Optimista N+2',                # AG
  34: f'Proy. Optimista {name_month_3}'   # AH
})

# ----- 3. Leemos Venta actual
wb_venta = load_workbook(filename_venta, data_only=True, read_only=True)
ws_venta = wb_venta['Venta - Plan']

for row in ws_venta.iter_rows(3, ws_venta.max_row, values_only=True):
  month_year = row[0]
  sector = row[1]
  oficina = row[3]
  material = row[4]
  descripcion = row[5]
  nivel_2 = row[6]
  venta_total = row[7] or 0
  plan_total = row[8] or 0
  canal_distribucion = 'Venta Directa'
  
  if oficina is not None:
    if oficina.lower() in dict_lead_time['optimista']['Venta Local'].keys():
      canal_distribucion = 'Venta Local'

    if int(venta_total) < 0:
      print(f'La venta de {oficina} {material} es negativa, correspondiente a: {venta_total}')
      venta_total = 0
    
    if int(plan_total) < 0:
      print(f'En plan de venta {oficina} {material} es negativo, correspondiente a: {plan_total}')
      plan_total = 0

    if month_year == selected_month1_year:
      ws.append({ 1: sector,
                  2: canal_distribucion,
                  3: f'{oficina.lower()}{material}',
                  4: oficina, 
                  5: int(material),
                  6: descripcion,
                  7: nivel_2, 
                  8: venta_total or 0,                # G
                  9: plan_total or 0,                 # H
                  10: 0,                              # J
                  11: 0,                              # K
                  20: 0
                })

wb_venta.close()

# ----- 4. Creamos la sheet Stock - Oficina
ws_stock_oficina = wb.create_sheet('Stock - Oficina')
stock(ws_stock_oficina, dict_lead_time, filename_dias)
wb.save(filename)

# ----- 5. Creamos la sheet Stock - ETA
print("--- %s 5.ETA inicio ---" % (time.time() - start_time))
ws_stock_ETA = wb.create_sheet(sheet_name_ETA)
dict_ETA_sin_venta = create_ETA(ws_stock_ETA, dict_lead_time, date_selected_month, dict_cierre_venta, dict_holidays, filename_logistica, filename_pedidos_confirmados)
wb.save(filename)
ETA_maxRow = ws_stock_ETA.max_row
print("--- %s 5. ---" % (time.time() - start_time))

# ----- 6. Agregamos Stock Puerto Oficina,	Almacen oficina
dict_stock = {}

for row in ws_stock_oficina.iter_rows(4, ws_stock_oficina.max_row, values_only=True):
  month_year = row[0]
  sector = row[1]
  oficina = row[2]
  material = row[3]
  descripcion = row[4]
  nivel_2 = row[5]
  llave = f'{oficina.lower()}{material}'
  # puerto_oficina = row[9] + row[13]
  # almacen = row[17] + row[21]
  # Stock solo liberado
  puerto_oficina = row[9]
  almacen = row[17]

  dict_stock[llave] = { 'sector': sector, 'oficina': oficina, 'material': material, 'descripcion': descripcion, 'nivel_2': nivel_2, 'Puerto oficina': puerto_oficina, 'Almacen': almacen }

# ----- 7. Producción faltante
wb_prod_faltante = load_workbook(filename_prod_faltante, read_only=True, data_only=True)
ws_prod_faltante = wb_prod_faltante.active
dict_prod_faltante = {}

for row in ws_prod_faltante.iter_rows(3, ws_prod_faltante.max_row, values_only=True):
  month_year = row[0]
  sector = row[1]
  material = int(row[2])
  descripcion = row[3]
  KG_plan = row[4]
  KG_real = row[5]

  if month_year not in dict_prod_faltante.keys():
    dict_prod_faltante[month_year] = {}

  dict_prod_faltante[month_year][material] = {
    'month_year': month_year,
    'sector': sector,
    'descripcion': descripcion,
    'KG_plan': KG_plan,
    'KG_real': KG_real
  }

wb_prod_faltante.close()

# ----- 8. RV FINAL
wb_RV = load_workbook(filename_RV, read_only=True, data_only=True)
ws_RV = wb_RV.active
dict_RV = {}

for row in ws_RV.iter_rows(2, ws_RV.max_row, values_only=True):
  month_year = row[0]
  sector = row[1]
  oficina = row[2]
  material = row[3]
  RV_final_prod = row[6]
  RV_final_venta = row[7]
  pedidos_ingresados = row[8]
  key = f'{oficina.lower()}{material}'
  
  if month_year not in dict_RV.keys():
    dict_RV[month_year] = {}
  
  if month_year in dict_RV.keys():
    dict_RV[month_year][key] = {
      'RV_final_prod': RV_final_prod,
      'RV_final_venta': RV_final_venta,
    }

wb_RV.close()

# ----- 9. Asignaciones de venta MES N+2
wb_asignaciones = load_workbook(filename_asignaciones, read_only=True, data_only=True)
ws_asignaciones = wb_asignaciones['Asignaciones de venta']
ws_asignaciones_max_row = ws_asignaciones.max_row
dict_asignaciones = {}
month_year = ''
sector = ''
office = ''

for row in ws_asignaciones.iter_rows(3, ws_asignaciones_max_row - 1, values_only=True):
  if row[0] is not None:
    month_year = str(row[0])
  
  if row[1] is not None:
    sector = row[1]

  if row[2] is not None:
    office = row[2]

  material = row[3]
  description = row[4]
  RV_final = row[6]

  key = f'{office.lower()}{material}'
  if month_year in dict_asignaciones:
    if key not in dict_asignaciones[month_year]:
      dict_asignaciones[month_year][key] = { 'sector': sector, 'oficina': oficina, 'material': material,'descripcion': description, 'RV final': RV_final }
  else:
    dict_asignaciones[month_year] = { key: { 'sector': sector, 'oficina': oficina, 'material': material,'descripcion': description, 'RV final': RV_final }}
wb_asignaciones.close()

# ----- 10. Creamos la sheet Stock - Puerto Chile
ws_puerto_chile = wb.create_sheet(sheet_name_PC)
create_puerto_chile(ws_puerto_chile, filename_chile, dict_lead_time, dict_holidays, month_1, dict_leftover_country, productive_days)
wb.save(filename)
PC_max_row = ws_puerto_chile.max_row

wb_agua = load_workbook(filename_chile, read_only=True, data_only=True)
ws_agua = wb_agua['Stock']
agua_max = ws_agua.max_row
dict_agua = {}

i = 4
for row in ws_agua.iter_rows(4, agua_max, values_only=True):
  month_year = row[0]
  sector = row[1]
  oficina = row[2]
  material = row[3]
  descripcion = row[4]
  nivel_2 = row[5]
  puerto_chile = row[8]
  key = f'{oficina.lower()}{material}'

  if month_year in dict_agua:
    if key not in dict_agua[month_year]:
      dict_agua[month_year][key] = {
        'fecha': month_year,
        'sector': sector,
        'oficina': oficina,
        'material': material,
        'descripcion': descripcion,
        'nivel 2': nivel_2,
        'puerto chile': puerto_chile
      }
  else:
    dict_agua[month_year] = {
      key: {
        'fecha': month_year,
        'sector': sector,
        'oficina': oficina,
        'material': material,
        'descripcion': descripcion,
        'nivel 2': nivel_2,
        'puerto chile': puerto_chile
      }
    }
  i += 1
wb_agua.close()

# ----- 11. STOCK CON VENTA Y PLAN
max_row = ws.max_row

print("--- %s 11. ---" % (time.time() - start_time))
for i, row in enumerate(ws.iter_rows(3, max_row, values_only = True), 3):
  canal_distribucion = row[1]
  llave = row[2]
  oficina = row[3]
  material = row[4]
  nivel_2 = row[6]
  lead_time_opt = dict_lead_time['optimista'][canal_distribucion][oficina.lower()]
  lead_time_pes = dict_lead_time['pesimista'][canal_distribucion][oficina.lower()]
  last_stop_LT = 0

  if canal_distribucion == 'Venta Directa':
    last_stop_LT = round(lead_time_opt['Puerto'], 2)
    leftover_days = dict_leftover_country['chile']
  else:
    last_stop_LT = round(lead_time_opt['Destino'], 2)
    leftover_days = dict_leftover_country[oficina.lower()]

  # -- 11.0. Chequeamos las ETAS sin venta
  if llave in dict_ETA_sin_venta.keys():
    dict_ETA_sin_venta.pop(llave)

  # -- 11.1. Prod Faltante
  if selected_month1_year in dict_prod_faltante:
    if int(material) in dict_prod_faltante[selected_month1_year]:
      ws[f'J{i}'].value = dict_prod_faltante[selected_month1_year][material]['KG_plan']
      ws[f'K{i}'].value = dict_prod_faltante[selected_month1_year][material]['KG_real']

  # -- 11.2. Plan - Prod. actual
  LT_opt_puerto = lead_time_opt['Puerto']
  LT_pes_puerto = lead_time_pes['Puerto']

  ws[f'L{i}'].value = f"=MAX(J{i} - K{i}, 0) * MAX(({leftover_days} - {LT_pes_puerto})/({LT_pes_puerto}), 0)"
  ws[f'R{i}'].value = f"=MAX(J{i} - K{i}, 0) * MAX(({leftover_days} - {LT_opt_puerto})/({LT_opt_puerto}), 0)"
    
  # -- 11.3.PUERTO CHILE mes 1
  # if selected_month1_year in dict_agua:
  #   if llave in dict_agua[selected_month1_year]:
  #     volumen_puerto_chile = dict_agua[selected_month1_year][llave]['puerto chile'] or 0
  #     ws[f'N{i}'].value = f"={volumen_puerto_chile} * ({leftover_days}/{productive_days})" or 0
  #     ws[f'T{i}'].value = f"={volumen_puerto_chile} * ({leftover_days}/{productive_days})" or 0
  #     dict_agua[selected_month1_year].pop(llave)

  ws[f'N{i}'].value = f"=SUMIF('{sheet_name_PC}'!$G$2:G{PC_max_row},'{sheet_name}'!C{i},'{sheet_name_PC}'!$L$2:L{PC_max_row})"
  ws[f'T{i}'].value = f"=SUMIF('{sheet_name_PC}'!$G$2:G{PC_max_row},'{sheet_name}'!C{i},'{sheet_name_PC}'!$N$2:N{PC_max_row})"
  
  # -- 11.4. Asignaciones mes 3
  if selected_month3_year in dict_asignaciones:
    if llave in dict_asignaciones[selected_month3_year]:
      ws[f'AD{i}'].value = dict_asignaciones[selected_month3_year][llave]['RV final'] or 0
      dict_asignaciones[selected_month3_year].pop(llave)

  # -- 11.5. MES N + 1
  # + ETAS
  # + inventario --> Proy mes N - Stock
  # - Ventas
  # Producción
  # -- Proyecciones mes N + 1
  if llave in dict_RV[selected_month2_year]:
    ws[f'X{i}'].value = dict_RV[selected_month2_year][llave]['RV_final_prod']          # Producción N + 1
    ws[f'Y{i}'].value = dict_RV[selected_month2_year][llave]['RV_final_venta']         # Venta N + 1

  ws[f'AA{i}'].value = f'=Z{i} + X{i}'                                          # PESIMISTA --> + ETA Pesimista n+1 + Producción N + 1
  ws[f'AC{i}'].value = f'=AB{i} + X{i}'                                         # OPTIMISTA --> + ETA Optimista n+1 + Producción N + 1
  ws[f'AA{i}'].fill = PatternFill("solid", fgColor=yellow)
  ws[f'AC{i}'].fill = PatternFill("solid", fgColor=yellow)

  # -- 11.6. MES N + 2
  porcentaje = dict_porcentaje_produccion[oficina.lower()]
  ws[f'AF{i}'].value = f'= {porcentaje} * AD{i} + AE{i}'                               # PESIMISTA --> Asignación de venta + ETA Pesimista n+2
  ws[f'AH{i}'].value = f'= {porcentaje} * AD{i} + AG{i}'                               # OPTIMISTA --> Asignación de venta + ETA Optimista n+2
  ws[f'AF{i}'].fill = PatternFill("solid", fgColor=yellow)
  ws[f'AH{i}'].fill = PatternFill("solid", fgColor=yellow)
  
  # VENTA LOCAL
  if canal_distribucion == "Venta Local":
    # -- ETA: Stock planta	Puerto Chile	Centro Agua
    ws[f'M{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$R$3:R{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$5)"
    ws[f'S{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$H$3:H{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$5)"

    ws[f'Z{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$S$3:S{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$R$3:R{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$7)"
    ws[f'AB{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$I$3:I{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$H$3:H{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$7)"
    
    ws[f'AE{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$T$3:T{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$S$3:S{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$8)"
    ws[f'AG{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$J$3:J{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$I$3:I{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$8)"

    # -- Stock Puerto Oficina y Almacen
    if llave in dict_stock.keys():
      stock_puerto_oficina = dict_stock[llave]['Puerto oficina'] or 0
      stock_almacen = dict_stock[llave]['Almacen'] or 0
      ws[f'O{i}'].value = f"={stock_puerto_oficina} * ({leftover_days} / {productive_days})"
      ws[f'P{i}'].value = f"={stock_almacen} * ({leftover_days} / {productive_days})"

      ws[f'U{i}'].value = f"={stock_puerto_oficina} * ({leftover_days} / {productive_days})"
      ws[f'V{i}'].value = f"={stock_almacen} * ({leftover_days} / {productive_days})"
      dict_stock.pop(llave, None)

    # -- Proyecciones mes N
    # + Venta Actual
    # + Stock Almacen
    # + Stock puerto
    # + Stock que va a llegar

    # Si no tengo venta
    # + RV prod
    # + Stock en el inventario
    # + Stock producido hoy
    # -Stock por despachar
    ws[f'Q{i}'].value = f'=H{i} + P{i} + MAX(M{i} - H{i} - P{i} + O{i},0)'             # PESIMISTA --> Venta Actual + Almacen oficina + ETA Pesimista n
    ws[f'Q{i}'].fill = PatternFill("solid", fgColor=yellow)
    ws[f'W{i}'].value = f'=H{i} + V{i} + MAX(S{i} - H{i} - V{i} - U{i},0)'             # OPTIMISTA --> Venta Actual + Almacen oficina + ETA Optimista n                                        # OPTIMISTA --> Venta Actual + Almacen oficina + ETA Optimista n
    ws[f'W{i}'].fill = PatternFill("solid", fgColor=yellow)
    
    if leftover_days >= last_stop_LT:
      ws[f'W{i}'].value = f'=H{i} + V{i} + U{i} + MAX(S{i} - H{i} - V{i} - U{i},0)'                                # OPTIMISTA --> + Puerto Oficina
      ws[f'W{i}'].fill = PatternFill("solid", fgColor=lightGreen)
  
  # VENTA DIRECTA
  elif canal_distribucion == "Venta Directa":                                          # VENTA EN PUERTO CHILE 
    ws[f'M{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$R$3:R{ETA_maxRow})"
    ws[f'S{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$H$3:H{ETA_maxRow})"

    ws[f'Z{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$S$3:S{ETA_maxRow})"
    ws[f'AB{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$I$3:I{ETA_maxRow})"

    ws[f'AE{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$T$3:T{ETA_maxRow})"
    ws[f'AG{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$J$3:J{ETA_maxRow})"

    # -- Proyecciones mes N
    if leftover_days > 0:
      LT_pes = lead_time_pes['Puerto']
      LT_opt = lead_time_opt['Puerto']
      pct_prod_pes = max(leftover_days - LT_pes, 0) / leftover_days
      pct_prod_opt = max(leftover_days - LT_opt, 0) / leftover_days

    ws[f'Q{i}'].value = f'=H{i} + {pct_prod_pes} * N{i} + MAX(M{i} - H{i} - ({pct_prod_pes} * N{i}), 0)'                   # PESIMISTA --> Venta Actual + ETA + Puerto Chile Pes. + (Plan - Prod. actual)
    ws[f'Q{i}'].fill = PatternFill("solid", fgColor=yellow)
    ws[f'W{i}'].value = f'=H{i} + {pct_prod_opt} * T{i} + MAX(S{i} - H{i} - ({pct_prod_pes} * T{i}), 0)'                   # OPTIMISTA --> Venta Actual + ETA + Puerto Chile Opt. + (Plan - Prod. actual)
    ws[f'W{i}'].fill = PatternFill("solid", fgColor=yellow)

  # ----- Styles
  thin = Side(border_style="thin", color=white)
  line_blue = Side(border_style="thin", color=blue)

  for j in range(8, ws.max_column + 1):
    ws[f'{get_column_letter(j)}{i}'].number_format = BUILTIN_FORMATS[3]
  
  ws[f'A{i}'].font = Font(bold=False, color=blue)
  ws[f'B{i}'].font = Font(bold=False, color=blue)
  ws[f'C{i}'].font = Font(bold=False, color=blue)
  ws[f'D{i}'].font = Font(bold=False, color=blue)
  ws[f'E{i}'].font = Font(bold=False, color=blue)
  ws[f'F{i}'].font = Font(bold=False, color=blue)
  ws[f'G{i}'].font = Font(bold=False, color=blue)
  ws[f'A{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'B{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'C{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'D{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'E{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'F{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'G{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'B{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'C{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'D{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'E{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'F{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'G{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

  # Linea separadora azul
  ws[f'H{i}'].border = Border(left=line_blue)
  ws[f'L{i}'].border = Border(left=line_blue)
  ws[f'X{i}'].border = Border(left=line_blue)
  ws[f'AD{i}'].border = Border(left=line_blue)
  ws[f'AI{i}'].border = Border(left=line_blue)

  # Bold optimista y pesimista
  ws[f'Q{i}'].font = Font(bold=True)
  ws[f'W{i}'].font = Font(bold=True)
  ws[f'Z{i}'].font = Font(bold=True)
  ws[f'AC{i}'].font = Font(bold=True)
  ws[f'AF{i}'].font = Font(bold=True)
  ws[f'AH{i}'].font = Font(bold=True)

# Merge 
ws.merge_cells('L1:W1')
ws.merge_cells('X1:AC1')
ws.merge_cells('AD1:AH1')

def add_line(ws, i, llave, value, canal_distribucion, dict_prod_faltante):
  prod_fat = dict_prod_faltante[selected_month1_year].get(material, {})

  ws.append({
    1: value['sector'],               # A
    2: canal_distribucion,            # B
    3: llave,                         # C
    4: value['oficina'],              # D
    5: value['material'],             # E
    6: value.get('descripcion', ''),  # F
    7: value.get('nivel_2', ''),      # G
    8: 0,                             # H
    9: 0,                             # I
    10: prod_fat.get('KG_plan', 0),   # J
    11: prod_fat.get('KG_real', 0),   # K
    
    15: value.get('Puerto oficina',0), # O
    16: value.get('Almacen', 0),       # P
    21: value.get('Puerto oficina', 0),# U
    22: value.get('Almacen', 0),       # V
  })

    # -- 12.1. Prod Faltante
  if selected_month1_year in dict_prod_faltante:
    if int(material) in dict_prod_faltante[selected_month1_year]:
      ws[f'J{i}'].value = dict_prod_faltante[selected_month1_year][material]['KG_plan']
      ws[f'K{i}'].value = dict_prod_faltante[selected_month1_year][material]['KG_real']
  
  # -- 12.2. Plan - Prod. actual
  LT_opt_puerto = lead_time_opt['Puerto']
  LT_pes_puerto = lead_time_pes['Puerto']

  ws[f'L{i}'].value = f"=MAX(J{i} - K{i}, 0) * MAX(({leftover_days} - {LT_pes_puerto})/({LT_pes_puerto}), 0)"
  ws[f'R{i}'].value = f"=MAX(J{i} - K{i}, 0) * MAX(({leftover_days} - {LT_opt_puerto})/({LT_opt_puerto}), 0)"
  
  # -- 12.3.PUERTO CHILE mes 1
  ws[f'N{i}'].value = f"=SUMIF('{sheet_name_PC}'!$G$2:G{PC_max_row},'{sheet_name}'!C{i},'{sheet_name_PC}'!$L$2:L{PC_max_row})"
  ws[f'T{i}'].value = f"=SUMIF('{sheet_name_PC}'!$G$2:G{PC_max_row},'{sheet_name}'!C{i},'{sheet_name_PC}'!$N$2:N{PC_max_row})"
  
  # -- 12.4. Asignaciones mes 3
  if selected_month3_year in dict_asignaciones:
    if llave in dict_asignaciones[selected_month3_year]:
      ws[f'AD{i}'].value = dict_asignaciones[selected_month3_year][llave]['RV final'] or 0
      dict_asignaciones[selected_month3_year].pop(llave)
  else:
    print(f'No se encuentra la fecha {selected_month3_year} en archivo {filename_asignaciones}')
  
  # -- 12.5. MES N + 1
  if llave in dict_RV[selected_month2_year]:
    ws[f'X{i}'].value = dict_RV[selected_month2_year][llave]['RV_final_prod']          # Producción N + 1
    ws[f'Y{i}'].value = dict_RV[selected_month2_year][llave]['RV_final_venta']         # Venta N + 1

  ws[f'AA{i}'].value = f'=Z{i} + X{i}'                                          # PESIMISTA --> + ETA Pesimista n+1 + Producción N + 1
  ws[f'AC{i}'].value = f'=AB{i} + X{i}'                                         # OPTIMISTA --> + ETA Optimista n+1 + Producción N + 1
  ws[f'AA{i}'].fill = PatternFill("solid", fgColor=yellow)
  ws[f'AC{i}'].fill = PatternFill("solid", fgColor=yellow)

  # -- 12.6. MES N + 2
  porcentaje = dict_porcentaje_produccion[oficina.lower()]
  ws[f'AF{i}'].value = f'= {porcentaje} * AD{i} + AE{i}'                               # PESIMISTA --> Asignación de venta + ETA Pesimista n+2
  ws[f'AH{i}'].value = f'= {porcentaje} * AD{i} + AG{i}'                               # OPTIMISTA --> Asignación de venta + ETA Optimista n+2
  ws[f'AF{i}'].fill = PatternFill("solid", fgColor=yellow)
  ws[f'AH{i}'].fill = PatternFill("solid", fgColor=yellow)

  # VENTA LOCAL
  if canal_distribucion == "Venta Local":
    # -- 12.7. ETA: Stock planta	Puerto Chile	Centro Agua
    ws[f'M{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$R$3:R{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$5)"
    ws[f'S{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$H$3:H{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$5)"

    ws[f'Z{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$S$3:S{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$R$3:R{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$7)"
    ws[f'AB{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$I$3:I{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$H$3:H{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$7)"
    
    ws[f'AE{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$T$3:T{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$S$3:S{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$AA$3:AA{ETA_maxRow},'{sheet_name}'!$AJ$8)"
    ws[f'AG{i}'].value = f"=SUMIFS('{sheet_name_ETA}'!$J$3:J{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$5) + SUMIFS('{sheet_name_ETA}'!$I$3:I{ETA_maxRow},'{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$Q$3:Q{ETA_maxRow},'{sheet_name}'!$AJ$8)"

    # -- 12.8. Stock Puerto Oficina y Almacen
    if llave in dict_stock.keys():
      stock_puerto_oficina = dict_stock[llave]['Puerto oficina'] or 0
      stock_almacen = dict_stock[llave]['Almacen'] or 0
      ws[f'O{i}'].value = f"={stock_puerto_oficina} * ({leftover_days} / {productive_days})"
      ws[f'P{i}'].value = f"={stock_almacen} * ({leftover_days} / {productive_days})"

      ws[f'U{i}'].value = f"={stock_puerto_oficina} * ({leftover_days} / {productive_days})"
      ws[f'V{i}'].value = f"={stock_almacen} * ({leftover_days} / {productive_days})"

    # -- 12.9. Proyecciones mes N
    ws[f'Q{i}'].value = f'=H{i} + N{i} + MAX(M{i} - H{i} - N{i}, 0)'                   # PESIMISTA --> Venta Actual + ETA + Puerto Chile Pes. + (Plan - Prod. actual)
    ws[f'Q{i}'].fill = PatternFill("solid", fgColor=yellow)
    ws[f'W{i}'].value = f'=H{i} + T{i} + MAX(S{i} - H{i} - T{i}, 0)'                   # OPTIMISTA --> Venta Actual + ETA + Puerto Chile Opt. + (Plan - Prod. actual)
    ws[f'W{i}'].fill = PatternFill("solid", fgColor=yellow)
    
    if leftover_days >= last_stop_LT:
      ws[f'W{i}'].value = f'=H{i} + V{i} + U{i} + MAX(S{i} - H{i} - V{i} - U{i},0)'                                # OPTIMISTA --> + Puerto Oficina
      ws[f'W{i}'].fill = PatternFill("solid", fgColor=lightGreen)
  
  # VENTA DIRECTA
  elif canal_distribucion == "Venta Directa":                                          # VENTA EN PUERTO CHILE 
    ws[f'M{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$R$3:R{ETA_maxRow})"
    ws[f'S{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$H$3:H{ETA_maxRow})"

    ws[f'Z{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$S$3:S{ETA_maxRow})"
    ws[f'AB{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$I$3:I{ETA_maxRow})"

    ws[f'AE{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$T$3:T{ETA_maxRow})"
    ws[f'AG{i}'].value = f"=SUMIF('{sheet_name_ETA}'!$F$3:F{ETA_maxRow},'{sheet_name}'!C{i},'{sheet_name_ETA}'!$J$3:J{ETA_maxRow})"

    # -- 12.10. Proyecciones mes N
    if leftover_days > 0:
      LT_pes = lead_time_pes['Puerto']
      LT_opt = lead_time_opt['Puerto']
      
      pct_prod_pes = max(leftover_days - LT_pes, 0) / leftover_days
      pct_prod_opt = max(leftover_days - LT_opt, 0) / leftover_days

    ws[f'Q{i}'].value = f'=H{i} + {pct_prod_pes} * N{i} + MAX(M{i} - H{i} - ({pct_prod_pes} * N{i}), 0)'                   # PESIMISTA --> Venta Actual + ETA + Puerto Chile Pes. + (Plan - Prod. actual)
    ws[f'Q{i}'].fill = PatternFill("solid", fgColor=yellow)
    ws[f'W{i}'].value = f'=H{i} + {pct_prod_opt} * T{i} + MAX(S{i} - H{i} - ({pct_prod_pes} * T{i}), 0)'                   # OPTIMISTA --> Venta Actual + ETA + Puerto Chile Opt. + (Plan - Prod. actual)
    ws[f'W{i}'].fill = PatternFill("solid", fgColor=yellow)
  
  # STYLES 
  thin = Side(border_style="thin", color=white)
  line_blue = Side(border_style="thin", color=blue)

  for j in range(8, ws.max_column + 1):
    ws[f'{get_column_letter(j)}{i}'].number_format = BUILTIN_FORMATS[3]

  ws[f'A{i}'].font = Font(bold=False, color=blue)
  ws[f'B{i}'].font = Font(bold=False, color=blue)
  ws[f'C{i}'].font = Font(bold=False, color=blue)
  ws[f'D{i}'].font = Font(bold=False, color=blue)
  ws[f'E{i}'].font = Font(bold=False, color=blue)
  ws[f'F{i}'].font = Font(bold=False, color=blue)
  ws[f'G{i}'].font = Font(bold=False, color=blue)
  ws[f'A{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'B{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'C{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'D{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'E{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'F{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'G{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
  ws[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'B{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'C{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'D{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'E{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'F{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws[f'G{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

  # Linea separadora azul
  ws[f'H{i}'].border = Border(left=line_blue)
  ws[f'L{i}'].border = Border(left=line_blue)
  ws[f'X{i}'].border = Border(left=line_blue)
  ws[f'AD{i}'].border = Border(left=line_blue)
  ws[f'AI{i}'].border = Border(left=line_blue)

  # Bold optimista y pesimista
  ws[f'Q{i}'].font = Font(bold=True)
  ws[f'W{i}'].font = Font(bold=True)
  ws[f'Z{i}'].font = Font(bold=True)
  ws[f'AC{i}'].font = Font(bold=True)
  ws[f'AF{i}'].font = Font(bold=True)
  ws[f'AH{i}'].font = Font(bold=True)

print("--- %s 12. ---" % (time.time() - start_time))
# ----- 12. STOCK SIN VENTA NI PLAN
for i, (key, value) in enumerate(dict_stock.copy().items(), ws.max_row + 1):
  oficina = value['oficina']
  material = value['material']
  llave = f'{oficina.lower()}{material}'
  porcentaje = dict_porcentaje_produccion[oficina.lower()]
  
  canal_distribucion = 'Venta Directa'
  leftover_days = dict_leftover_country['chile']
  last_stop_LT = round(lead_time_opt['Puerto'], 2)

  if oficina.lower() in dict_lead_time['optimista']['Venta Local'].keys():
    canal_distribucion = 'Venta Local'
    leftover_days = dict_leftover_country[oficina.lower()]

  
  lead_time_opt = dict_lead_time['optimista'][canal_distribucion][oficina.lower()]
  lead_time_pes = dict_lead_time['pesimista'][canal_distribucion][oficina.lower()]
  last_stop_LT = round(lead_time_opt['Destino'], 2)
  
  if llave in dict_ETA_sin_venta.keys():
    dict_ETA_sin_venta.pop(llave)
  
  add_line(ws, i, llave, value, canal_distribucion, dict_prod_faltante)

# ----- 13. ETAS SIN VENTA
# for i, (key, value) in enumerate(dict_ETA_sin_venta.items(), ws.max_row + 1):
#   canal_distribucion = value['canal_distribucion']
#   oficina = value['oficina']
#   material = value['material']
#   sector = value['sector']

#   add_line(ws, i, key, value, canal_distribucion, dict_prod_faltante)



# ----- 13. Guardar la información
run_styles(ws)

ws['AJ5'].value = "SI"
ws['AJ6'].value = f"Mes {month_1.month}"
ws['AJ7'].value = f"Mes {month_2.month}"
ws['AJ8'].value = f"Mes {month_3.month}"

ws['AJ3'].fill = PatternFill("solid", fgColor=yellow)
ws['AK3'].value = 'Se suma los pedidos de puerto oficina'

ws['AJ4'].fill = PatternFill("solid", fgColor=lightGreen)

ws['AK4'].value = 'Se suma los pedidos que llegan este mes de agua'
print("--- %s 11. ---" % (time.time() - start_time))

wb.save(filename)
wb.close()
print("--- %s seconds ---" % (time.time() - start_time))
# messageBox(dict_lead_time, 'Venta Local', filename_dias, path_img)