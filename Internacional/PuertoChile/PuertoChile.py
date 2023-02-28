from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS, FORMAT_PERCENTAGE
from datetime import datetime, date
import calendar
from constants import *
from openpyxl.utils import get_column_letter

def create_puerto_chile(ws, filename_chile, dict_lead_time, dict_holidays, month_1, dict_leftover_country, productive_days):
  name_month_1 = month_translate_EN_CL[month_1.strftime('%B').lower()]
  today = datetime.now().date()

  ws.append({
    1: 'Fecha',               # A
    2: 'Sector',              # B
    3: 'Oficina',             # C
    4: 'Material',            # D
    5: 'Descripción',         # E
    6: 'Nivel 2',             # F
    7: 'Llave',               # G
    8: 'Puerto Chile',
    9: f'Días productivos desde\n{today}',
    10: f'Días productivos mensual {name_month_1}',
    11: 'Porcentaje prod. Pes.',
    12: 'Stock vendible Pes.',
    13: 'Porcentaje prod Opt.',
    14: 'Stock vendible Opt.'
  })

  thin = Side(border_style="thin", color=white)
  for i in range(1, ws.max_column + 1):
    ws[f'{get_column_letter(i)}1'].font = Font(bold=True, color=white)
    ws[f'{get_column_letter(i)}1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'{get_column_letter(i)}1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{get_column_letter(i)}1'].fill = PatternFill("solid", fgColor=lightBlue)

  wb_agua = load_workbook(filename_chile, read_only=True, data_only=True)
  ws_agua = wb_agua['Stock']
  agua_max = ws_agua.max_row

  i = 2
  for row in ws_agua.iter_rows(4, agua_max, values_only=True):
    month_year = row[0]
    sector = row[1]
    oficina = row[2]
    material = int(row[3])
    descripcion = row[4]
    nivel_2 = row[5]
    puerto_chile = row[8] or 0
    key = f'{oficina.lower()}{material}'

    canal_distribucion = 'Venta Directa'
    holidays_country = dict_holidays[month_1.year]['chile']
    leftover_days = dict_leftover_country['chile']
  
    if oficina is not None:
      if oficina.lower() in dict_lead_time['optimista']['Venta Local'].keys():
        canal_distribucion = 'Venta Local'
        holidays_country = dict_holidays[month_1.year][oficina.lower()]
        leftover_days = dict_leftover_country[oficina.lower()]

    if oficina.lower() in dict_lead_time['optimista'][canal_distribucion]:
      lead_time_opt = dict_lead_time['optimista'][canal_distribucion][oficina.lower()]
      lead_time_pes = dict_lead_time['pesimista'][canal_distribucion][oficina.lower()]
      
    LT_pes = lead_time_pes['Puerto']
    LT_opt = lead_time_opt['Puerto']
    today = datetime.now()
    last_day_month = calendar.monthrange(today.year, today.month)[1]

    # Porcentaje producción
    if leftover_days > 0:
      pct_prod_pes = max(leftover_days - LT_pes, 0) / leftover_days
      pct_prod_opt = max(leftover_days - LT_opt, 0) / leftover_days
    
    pct_prod_opt = 0
    pct_prod_pes = 0

    ws[f'A{i}'].value = month_year
    ws[f'B{i}'].value = sector
    ws[f'C{i}'].value = oficina
    ws[f'D{i}'].value = material
    ws[f'E{i}'].value = descripcion
    ws[f'F{i}'].value = nivel_2
    ws[f'G{i}'].value = key
    ws[f'H{i}'].value = puerto_chile
    ws[f'I{i}'].value = leftover_days
    ws[f'J{i}'].value = productive_days
    ws[f'K{i}'].value = pct_prod_pes
    ws[f'L{i}'].value = f"=J{i} * K{i}"
    ws[f'M{i}'].value = pct_prod_opt
    ws[f'N{i}'].value = f"=M{i} * H{i}"

    # ----- Styles -----
    thin = Side(border_style="thin", color=white)
    line_blue = Side(border_style="thin", color=blue)

    ws[f'A{i}'].font = Font(bold=False, color=blue)
    ws[f'A{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    ws[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws[f'B{i}'].font = Font(bold=False, color=blue)
    ws[f'B{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    ws[f'B{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    ws[f'C{i}'].font = Font(bold=False, color=blue)
    ws[f'C{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    ws[f'C{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws[f'D{i}'].font = Font(bold=False, color=blue)
    ws[f'D{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    ws[f'D{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws[f'E{i}'].font = Font(bold=False, color=blue)
    ws[f'E{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    ws[f'E{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws[f'F{i}'].font = Font(bold=False, color=blue)
    ws[f'F{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    ws[f'F{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws[f'G{i}'].font = Font(bold=False, color=blue)
    ws[f'G{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    ws[f'G{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Tamaños
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 23
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['M'].width = 10

    # Styles
    ws[f'H{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'K{i}'].number_format = FORMAT_PERCENTAGE
    ws[f'M{i}'].number_format = FORMAT_PERCENTAGE
    i += 1

