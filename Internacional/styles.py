
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, BUILTIN_FORMATS
from constants import *

import time
start_time = time.time()

def run_styles(ws):
  thin = Side(border_style="thin", color=white)
  line_blue = Side(border_style="thin", color=blue)

  for letter in ['L', 'X', 'AD']:
    ws[f'{letter}1'].font = Font(bold=True, color=white)
    ws[f'{letter}1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'{letter}1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{letter}1'].fill = PatternFill("solid", fgColor=lightBlue)

  for i in range(1, ws.max_column):
    ws[f'{get_column_letter(i)}2'].font = Font(bold=True, color=white)
    ws[f'{get_column_letter(i)}2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'{get_column_letter(i)}2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{get_column_letter(i)}2'].fill = PatternFill("solid", fgColor=lightBlue)

  ws['Q2'].fill = PatternFill("solid", fgColor=blue)
  ws['W2'].fill = PatternFill("solid", fgColor=blue)
  ws['Z2'].fill = PatternFill("solid", fgColor=blue)
  ws['AC2'].fill = PatternFill("solid", fgColor=blue)
  ws['AF2'].fill = PatternFill("solid", fgColor=blue)
  ws['AH2'].fill = PatternFill("solid", fgColor=blue)

  # Tamaños
  ws.column_dimensions['B'].width = 11
  ws.column_dimensions['C'].width = 23
  ws.column_dimensions['D'].width = 16
  ws.column_dimensions['F'].width = 32
  ws.column_dimensions['G'].width = 16

  for i in range(8, ws.max_column + 1):
    ws.column_dimensions[f'{get_column_letter(i)}'].width = 11

  # Tamaño
  ws.row_dimensions[1].height =  25