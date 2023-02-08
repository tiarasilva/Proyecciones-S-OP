from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, BUILTIN_FORMATS
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from Stock.stock import stock
from ETA.ETA_VL import create_ETA_VL
from ETA.ETA_VD import create_ETA_VD
from MessageBox.MessageBox import messageBox
from styles import run_styles, run_number_format
from constants import *

import time
import calendar
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import holidays
# from TD.pivotTable import pivot_table

import sys
import os

start_time = time.time()

# ----- 0. PATH
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
  print('running in a PyInstaller bundle')
  os.chdir(sys._MEIPASS)
else:
  print('running in a normal Python process')

# ----- 1. Abrimos el excel de los parametros
wb_parametros = load_workbook(filename_parametros, data_only = True, read_only = True)
ws_parametros_time = wb_parametros['Lead time']
ws_parametros_venta = wb_parametros['Venta']
ws_porcentaje = wb_parametros['Asignación productivo']

# DIAS PARA VENDER
ws_parametros_dias_venta = wb_parametros['Cierre venta']
dict_cierre_venta = {}
ws_dias_max = ws_parametros_dias_venta.max_row

for row in ws_parametros_dias_venta.iter_rows(2, ws_dias_max, values_only = True):
  if row[1] is None:
    break
  oficina = row[1]
  dias_cierre = row[2]
  dict_cierre_venta[oficina.lower()] = dias_cierre

# FECHAS
selected_month = ws_parametros_venta['B1'].value
selected_week = ws_parametros_venta['B2'].value
number_selected_month = month_number[selected_month.lower()]
today = datetime.now()
date_selected_month = date(today.year, number_selected_month, 1)
last_day_month = calendar.monthrange(today.year, today.month)[1]

# HOLIDAYS
last_year = today.year - 1
this_year = today.year
next_year = today.year + 1

dict_holidays = {
  last_year: {
    'agro america': holidays.US(years=last_year),         # USA
    'agro europa': holidays.IT(years=last_year),          # ITALIA
    'agro mexico': holidays.MX(years=last_year),          # Mexico
    'agrosuper shanghai': holidays.CN(years=last_year),   # China
    'andes asia': holidays.KR(years=last_year),           # Corea del sur
    'chile': holidays.CL(years=last_year)
  },
  today.year: {
    'agro america': holidays.US(years=this_year),         # USA
    'agro europa': holidays.IT(years=this_year),          # ITALIA
    'agro mexico': holidays.MX(years=this_year),          # Mexico
    'agrosuper shanghai': holidays.CN(years=this_year),   # China
    'andes asia': holidays.KR(years=this_year),           # Corea del sur
    'chile': holidays.CL(years=this_year)
  },
  next_year: {
    'agro america': holidays.US(years=next_year),         # USA
    'agro europa': holidays.IT(years=next_year),          # ITALIA
    'agro mexico': holidays.MX(years=next_year),          # Mexico
    'agrosuper shanghai': holidays.CN(years=next_year),   # China
    'andes asia': holidays.KR(years=next_year),           # Corea del sur
    'chile': holidays.CL(years=next_year)
  }
}

# TIPO DE VENTA y LEAD TIME
venta_iteracion = 'directa'
escenario_iteracion = 'optimista'
dict_lead_time = {
  'optimista': {'directa': {}, 'local': {}},
  'pesimista': {'directa': {}, 'local': {}},
}

for row in ws_parametros_time.iter_rows(2, ws_parametros_time.max_row, values_only = True):
  if 'Venta Local' == row[0]:
    venta_iteracion = 'local'
  
  if 'PESIMISTA' == row[0]:
    escenario_iteracion = 'pesimista'
    venta_iteracion = 'directa'

  oficina = row[1]
  planta = row[2]
  puerto = row[3]
  agua = row[4]
  destino = row[5]
  almacen = row[6]

  if oficina is not None:
    if oficina.lower() != 'oficina':
      if 'local' in venta_iteracion.lower():
        dict_lead_time[escenario_iteracion][venta_iteracion][oficina.lower()] = { 'Planta': planta, 'Puerto': puerto, 'Agua': agua, 'Destino': destino, 'Almacen': almacen }
      else:
        dict_lead_time[escenario_iteracion][venta_iteracion][oficina.lower()] = { 'Planta': planta, 'Puerto': puerto }

# PORCENTAJE DE PRODUCCIÓN
dict_porcentaje_produccion = {}

for row in ws_porcentaje.iter_rows(2, ws_porcentaje.max_row, values_only=True):
  oficina = row[1].lower()
  produccion = row[2]
  dict_porcentaje_produccion[oficina] = produccion

wb_parametros.close()

# ----. Nombre fechas
month_1 = date_selected_month
month_2 = date_selected_month + relativedelta(months=1)
month_3 = date_selected_month + relativedelta(months=2)

name_month_1 = month_translate_EN_CL[month_1.strftime('%B').lower()]
name_month_2 = month_translate_EN_CL[month_2.strftime('%B').lower()]
name_month_3 = month_translate_EN_CL[month_3.strftime('%B').lower()]

# ----- 2. Creamos el excel de resultados VENTA LOCAL
wb_VL = Workbook()
ws_VL = wb_VL.active
ws_VL.title = sheet_name

ws_VL.append({
  8: f'Proyección {name_month_1} {month_1.year}',
  14: f'Proyección {name_month_2} {month_2.year}',
  18: f'Proyección {name_month_3} {month_3.year}'
})

ws_VL.append({
  1: 'Sector',                # A
  2: 'Llave',                 # B
  3: 'Oficina',               # C
  4: 'Material',              # D
  5: 'Descripción',           # E
  6: 'Venta Actual',          # F
  7: 'Plan',                  # G
  8: 'ETA Pesimista',         # H
  9: 'ETA Optimista',         # I
  10: 'Puerto Oficina',       # J
  11: 'Almacen oficina',      # Q
  12: 'Pesimista Proy.',      # L
  13: 'Optimista Proy.',      # M
  14: 'ETA Pesimista',        # N
  15: 'ETA Optimista',        # O
  16: 'Pesimista Proy.2',     # P
  17: 'Optimista Proy.2',     # Q
  18: 'Asignación de venta',  # R
  19: 'ETA Pesimista',        # S
  20: 'ETA Optimista',        # T
  21: 'Pesimista Proy.3',     # U
  22: 'Optimista Proy.3'      # V
})

# ----- y VENTA DIRECTA
wb_VD = Workbook()
ws_VD = wb_VD.active
ws_VD.title = sheet_name

ws_VD.append({
  8: f'Proyección {name_month_1} {month_1.year}',
  12: f'Proyección {name_month_2} {month_2.year}',
  16: f'Proyección {name_month_3} {month_3.year}'
})

ws_VD.append({
  1: 'Sector',                # A
  2: 'Llave',                 # B
  3: 'Oficina',               # C
  4: 'Material',              # D
  5: 'Descripción',           # E
  6: 'Venta Actual',          # F
  7: 'Plan',                  # G
  8: 'ETA Pesimista',         # H
  9: 'ETA Optimista',         # I
  10: 'Pesimista Proy.',      # J
  11: 'Optimista Proy.',      # K
  12: 'ETA Pesimista',        # L
  13: 'ETA Optimista',        # M
  14: 'Pesimista Proy.2',     # N
  15: 'Optimista Proy.2',     # O
  16: 'Asignación de venta',  # P
  17: 'ETA Pesimista',        # Q
  18: 'ETA Optimista',        # R
  19: 'Pesimista Proy.3',     # S
  20: 'Optimista Proy.3'      # T
})

# ----- 3. Leemos Venta actual
wb_venta = load_workbook(filename_venta, data_only=True, read_only=True)
ws_venta = wb_venta['Venta - Plan']

for row in ws_venta.iter_rows(7, ws_venta.max_row, values_only=True):
  sector, material = row[1], row[2]
  descripcion, oficina = row[3], row[5]
  plan_total, venta_total = row[6], row[7]
  
  if sector is not None and descripcion is not None:
    ws = ws_VD
    if oficina.lower() in dict_lead_time['optimista']['local'].keys():
      ws = ws_VL

    ws.append({ 1: sector,
                2: f'{oficina.lower()}{material}',
                3: oficina, 
                4: int(material), 
                5: descripcion, 
                6: venta_total or 0, 
                7: plan_total or 0,
                8: 0,
              })

wb_venta.close()

# ----- 4. Creamos la sheet Stock - Oficina
ws_VL_stock_oficina = wb_VL.create_sheet('Stock - Oficina')
stock(ws_VL_stock_oficina, dict_lead_time, 'local', selected_month)
wb_VL.save(filename_VL)

# ----- 5. Creamos la sheet Stock - Oficina
print("--- %s 4.ETA inicio ---" % (time.time() - start_time))
ws_VL_stock_ETA = wb_VL.create_sheet(sheet_stock)
create_ETA_VL(ws_VL_stock_ETA, dict_lead_time, date_selected_month, dict_cierre_venta, dict_holidays)
wb_VL.save(filename_VL)
ETA_maxRow_VL = ws_VL_stock_ETA.max_row

print("--- %s 5. ETA final---" % (time.time() - start_time))

ws_VD_stock_ETA = wb_VD.create_sheet(sheet_stock)
create_ETA_VD(ws_VD_stock_ETA, dict_lead_time, date_selected_month, dict_cierre_venta, dict_holidays)
wb_VD.save(filename_VD)
ETA_maxRow_VD = ws_VD_stock_ETA.max_row
print("--- %s 6. ETA final---" % (time.time() - start_time))

# ----- 6. Agregamos Stock Puerto Oficina,	Almacen oficina
dict_stock = {  }

for row in ws_VL_stock_oficina.iter_rows(4, ws_VL_stock_oficina.max_row, values_only=True):
  sector = row[0]
  oficina = row[1]
  material = row[2]
  descripcion = row[3]
  llave = f'{oficina.lower()}{material}'
  puerto_oficina = row[8] + row[12]
  almacen = row[16] + row[20]
  dict_stock[llave] = { 'sector': sector, 'oficina': oficina, 'material': material, 'descripcion': descripcion, 'Puerto oficina': puerto_oficina, 'Almacen': almacen }

print("--- %s 7. ---" % (time.time() - start_time))

# ----- 7. Asignaciones de venta MES N+2
wb_asignaciones = load_workbook(filename_asignaciones, read_only=True, data_only=True)
ws_asignaciones = wb_asignaciones['Asignaciones de venta']
ws_asignaciones_max_row = ws_asignaciones.max_row
dict_asignaciones = {}
month_year = ''
sector = ''
office = ''

for row in ws_asignaciones.iter_rows(3, ws_asignaciones_max_row - 1, values_only=True):
  if row[0] is not None:
    month_year = str(row[0])
  
  if row[1] is not None:
    sector = row[1]

  if row[2] is not None:
    office = row[2]

  material = row[3]
  description = row[4]
  RV_final = row[6]

  key = f'{office.lower()}{material}'
  if month_year in dict_asignaciones:
    if key not in dict_asignaciones[month_year]:
      dict_asignaciones[month_year][key] = { 'sector': sector, 'oficina': oficina, 'material': material,'descripcion': description, 'RV final': RV_final }
  else:
    dict_asignaciones[month_year] = { key: { 'sector': sector, 'oficina': oficina, 'material': material,'descripcion': description, 'RV final': RV_final }}
wb_asignaciones.close()

# -----
dict_leftover_country = {}

key_month_year = f'{month_3.month}.{month_3.year}'
if month_3.month < 10:
  key_month_year = f'0{month_3.month}.{month_3.year}'

# VENTA LOCAL
for i, row in enumerate(ws_VL.iter_rows(3, ws_VL.max_row, values_only = True), 3):
  llave = row[1]
  oficina = row[2]
  lead_time_opt = dict_lead_time['optimista']['local'][oficina.lower()]
  holidays_country = dict_holidays[month_1.year][oficina.lower()]

  if llave in dict_stock:
    ws_VL[f'J{i}'].value = dict_stock[llave]['Puerto oficina'] or 0
    ws_VL[f'K{i}'].value = dict_stock[llave]['Almacen'] or 0
    dict_stock.pop(llave, None)

  if key_month_year in dict_asignaciones:
    if llave in dict_asignaciones[key_month_year]:
      ws_VL[f'R{i}'].value = dict_asignaciones[key_month_year][llave]['RV final']
      dict_asignaciones[key_month_year].pop(llave)

  # -- Feriados
  leftover_days = 0
  if oficina in dict_leftover_country:
    leftover_days = dict_leftover_country[oficina.lower()]
  else:
    for day in range(today.day + 1, last_day_month + 1):
      date_day = date(today.year, number_selected_month, day)
      if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
        leftover_days += 1
    dict_leftover_country[oficina.lower()] = leftover_days
  
  # -- MES N
  ws_VL[f'L{i}'].value = f'=F{i} + K{i} + H{i}'                    # PESIMISTA --> Venta Actual + Almacen oficina
  ws_VL[f'M{i}'].value = f'=F{i} + K{i} + I{i}'                    # OPTIMISTA --> Venta Actual + Almacen oficina
  ws_VL[f'M{i}'].fill = PatternFill("solid", fgColor=yellow)

  if leftover_days >= round(lead_time_opt['Destino'], 2):
    ws_VL[f'M{i}'].value = f'=F{i} + K{i} + J{i} + I{i}'           # OPTIMISTA --> + Puerto Oficina
    ws_VL[f'M{i}'].fill = PatternFill("solid", fgColor=lightGreen)
  
  # -- MES N + 1
  # + no alcance a vender del MES N
  # + producción de este mes
  # + ETAS
  ws_VL[f'P{i}'].value = f'=N{i}'
  ws_VL[f'Q{i}'].value = f'=O{i}'

  # -- MES N + 2
  # + no alcance a vender del MES N + 1
  # + 0.7 * Asignación de venta
  # + ETAS
  porcentaje = dict_porcentaje_produccion[oficina.lower()]
  ws_VL[f'U{i}'].value = f'= {porcentaje} * R{i} + S{i}'           # PESIMISTA --> Asignación de venta + ETA Pesimista n+2
  ws_VL[f'V{i}'].value = f'= {porcentaje} * R{i} + T{i}'           # OPTIMISTA --> Asignación de venta + ETA Optimista n+2

    # ----- Stock planta	Puerto Chile	Centro Agua
  ws_VL[f'I{i}'].value = f"=SUMIFS('{sheet_stock}'!$G$3:G{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$5)"
  ws_VL[f'O{i}'].value = f"=SUMIFS('{sheet_stock}'!$H$3:H{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$G$3:G{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$7)"
  ws_VL[f'T{i}'].value = f"=SUMIFS('{sheet_stock}'!$I$3:I{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$H$3:H{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$8)"

  ws_VL[f'H{i}'].value = f"=SUMIFS('{sheet_stock}'!$Q$3:Q{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$5)"
  ws_VL[f'N{i}'].value = f"=SUMIFS('{sheet_stock}'!$R$3:R{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$Q$3:Q{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$7)"
  ws_VL[f'S{i}'].value = f"=SUMIFS('{sheet_stock}'!$S$3:S{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$R$3:R{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$8)"

# VENTA DIRECTA
for i, row in enumerate(ws_VD.iter_rows(3, ws_VD.max_row, values_only = True), 3):
  llave = row[1]
  oficina = row[2]
  lead_time_opt = dict_lead_time['optimista']['directa'][oficina.lower()]
  holidays_country = dict_holidays[month_1.year]['chile']

  if key_month_year in dict_asignaciones:
    if llave in dict_asignaciones[key_month_year]:
      ws_VD[f'P{i}'].value = dict_asignaciones[key_month_year][llave]['RV final']
      dict_asignaciones[key_month_year].pop(llave)

  # -- Feriados
  leftover_days = 0
  if oficina in dict_leftover_country:
    leftover_days = dict_leftover_country[oficina.lower()]
  else:
    for day in range(today.day + 1, last_day_month + 1):
      date_day = date(today.year, number_selected_month, day)
      if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
        leftover_days += 1
    dict_leftover_country[oficina.lower()] = leftover_days
  
  # -- MES N
  ws_VD[f'J{i}'].value = f'=F{i} + H{i}'                    # PESIMISTA --> Venta Actual + ETA Pesimista + Puerto Oficina
  ws_VD[f'K{i}'].value = f'=F{i} + I{i}'                    # OPTIMISTA --> Venta Actual + ETA Optimista + Puerto Oficina
  
  # -- MES N + 1
  ws_VD[f'N{i}'].value = f'=L{i}'
  ws_VD[f'O{i}'].value = f'=M{i}'                                  # OPTIMISTA --> ETAS

  # -- MES N + 2
  porcentaje = dict_porcentaje_produccion[oficina.lower()]
  ws_VD[f'S{i}'].value = f'= {porcentaje} * P{i} + Q{i}'           # PESIMISTA --> Asignación de venta + ETA Pesimista n+2
  ws_VD[f'T{i}'].value = f'= {porcentaje} * P{i} + R{i}'           # OPTIMISTA --> Asignación de venta + ETA Optimista n+2

  ws_VD[f'I{i}'].value = f"=SUMIFS('{sheet_stock}'!$G$2:G{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$L$2:L{ETA_maxRow_VD},'{sheet_name}'!$X$5)"
  ws_VD[f'M{i}'].value = f"=SUMIFS('{sheet_stock}'!$H$2:H{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$L$2:L{ETA_maxRow_VD},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$G$2:G{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$L$2:L{ETA_maxRow_VD},'{sheet_name}'!$X$7)"
  ws_VD[f'R{i}'].value = f"=SUMIFS('{sheet_stock}'!$I$2:I{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$L$2:L{ETA_maxRow_VD},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$H$2:H{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$L$2:L{ETA_maxRow_VD},'{sheet_name}'!$X$8)"

  # ws_VD[f'H{i}'].value = f"=SUMIFS('{sheet_stock}'!$Q$2:Q{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$2:Z{ETA_maxRow_VD},'{sheet_name}'!$X$5)"
  # ws_VD[f'L{i}'].value = f"=SUMIFS('{sheet_stock}'!$R$2:R{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$2:Z{ETA_maxRow_VD},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$Q$2:Q{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$2:Z{ETA_maxRow_VD},'{sheet_name}'!$X$7)"
  # ws_VD[f'Q{i}'].value = f"=SUMIFS('{sheet_stock}'!$S$2:S{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$2:Z{ETA_maxRow_VD},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$R$2:R{ETA_maxRow_VD},'{sheet_stock}'!$E$2:E{ETA_maxRow_VD},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$2:Z{ETA_maxRow_VD},'{sheet_name}'!$X$8)"

print("--- %s 8. ---" % (time.time() - start_time))
# ----- Stock sin Venta ni Plan
i = ws_VL.max_row

for key, value in dict_stock.items():
  of = value['oficina']
  mat = value['material']
  porcentaje = dict_porcentaje_produccion[of.lower()]

  ws_VL.append({
    1: dict_stock[key]['sector'],
    2: f'{of}{mat}',
    3: of,
    4: mat,
    5: dict_stock[key]['descripcion'],
    6: 0,
    7: 0,
    8: 0
  })
  
  if oficina.lower() in dict_lead_time['optimista']['local']:
    i += 1
    ws[f'I{i}'].value = f"=SUMIFS('{sheet_stock}'!$G$3:G{ETA_maxRow_VL}, '{sheet_stock}'!$E$3:E{ETA_maxRow_VL}, '{sheet_name}'!B{i}, '{sheet_stock}'!$P$3:P{ETA_maxRow_VL}, '{sheet_name}'!$X$5)"
    ws[f'J{i}'].value = f'=F{i} + H{i} + J{i}'
    ws[f'K{i}'].value = f'=F{i} + I{i} + J{i}'
    ws[f'L{i}'].value = f"=SUMIFS('{sheet_stock}'!$R$3:R{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$Q$3:Q{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$7)"
    ws[f'M{i}'].value = f"=SUMIFS('{sheet_stock}'!$H$3:H{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$G$3:G{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$7)"
    ws[f'N{i}'].value = f'=L{i}'
    ws[f'O{i}'].value = f'=M{i}' 
    ws[f'P{i}'].value = dict_asignaciones[key_month_year][llave]['RV final'] or 0
    ws[f'Q{i}'].value = f"=SUMIFS('{sheet_stock}'!$S$3:S{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$R$3:R{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$Z$3:Z{ETA_maxRow_VL},'{sheet_name}'!$X$8)"
    ws[f'R{i}'].value = f"=SUMIFS('{sheet_stock}'!$I$3:I{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$5) + SUMIFS('{sheet_stock}'!$H$3:H{ETA_maxRow_VL},'{sheet_stock}'!$E$3:E{ETA_maxRow_VL},'{sheet_name}'!B{i},'{sheet_stock}'!$P$3:P{ETA_maxRow_VL},'{sheet_name}'!$X$8)"
    ws[f'S{i}'].value = f'= {porcentaje} * P{i} + Q{i}'
    ws[f'T{i}'].value = f'= {porcentaje} * P{i} + R{i}'
    
  
print("--- %s 9. ---" % (time.time() - start_time))

# ----- Guardar la información
run_styles(ws_VL, 'local')
run_styles(ws_VD, 'di')

for ws in [ws_VD, ws_VL]:
  run_number_format(ws)

  ws['X5'].value = "SI"
  ws['X6'].value = f"Mes {month_1.month}"
  ws['X7'].value = f"Mes {month_2.month}"
  ws['X8'].value = f"Mes {month_3.month}"

  ws['X3'].fill = PatternFill("solid", fgColor=yellow)
  ws['Y3'].value = 'Se suma los pedidos de puerto oficina'

  ws['X4'].fill = PatternFill("solid", fgColor=lightGreen)
  ws['Y4'].value = 'Se suma los pedidos que llegan este mes de agua'
print("--- %s 11. ---" % (time.time() - start_time))

wb_VL.save(filename_VL)
print(wb_VL.sheetnames)
print(wb_VL['Rango proyecciones'].max_row)
wb_VL.close()

wb_VD.save(filename_VD)
wb_VD.close()

# ----- Sheet TD
# main()
# pivot_table()

# user_response = input('Desea chequear la proyección?: (Si/No)')

# if user_response in ['Si', 'si', 'SI']:
#   wb_proy = load_workbook()

print("--- %s seconds ---" % (time.time() - start_time))
messageBox(dict_lead_time, 'local')