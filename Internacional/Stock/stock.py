from openpyxl import load_workbook
from Stock.styles import *
from constants import *

import calendar
from datetime import datetime, date
import holidays

from os import path

# ----- Creamos el sheet
def stock(ws, dict_lead_time, filename_dias):
  dict_message = {'Puerto Oficina': {}, 'Almacen': {}}

  ws.append({
    7: 'Puerto Oficina', 
    15: 'Almacén Oficina'
  })

  ws.append({
    7: 'Stock liberado', 11: 'Stock no liberado',    # Puerto Oficina
    15: 'Stock liberado', 19: 'Stock no liberado',   # Almacen Oficina
  })
  
  ws.append({
    1: 'Fecha',
    2: 'Sector', 
    3: 'Oficina', 
    4: 'Material', 
    5: 'Descripción', 
    6: 'Nivel 2', 
    7: 'Total KG', 8: 'Nº Días de Antigüedad Centro', 9: 'Nº Días de Antigüedad Oficina', 10: 'Status considerado',
    11: 'Total KG', 12: 'Nº Días de Antigüedad Centro', 13: 'Nº Días de Antigüedad Oficina', 14: 'Status considerado',
    15: 'Total KG', 16: 'Nº Días de Antigüedad Centro', 17: 'Nº Días de Antigüedad Oficina', 18: 'Status considerado',
    19: 'Total KG', 20: 'Nº Días de Antigüedad Centro', 21: 'Nº Días de Antigüedad Oficina', 22: 'Status considerado',
  })

  # ----- Días antiguedad Stock
  wb_dias_stock = load_workbook(filename_dias, read_only=True, data_only=True)
  ws_dias_stock = wb_dias_stock['Stock']
  dict_dias_stock = {}
  obj = calendar.Calendar()

  sector = ''
  oficina = ''
  i = 4
  for row in ws_dias_stock.iter_rows(5, ws_dias_stock.max_row - 1, values_only=True):
    if row[3] is None:
      break
    
    if row[1] is not None:
      sector = row[1]
    
    if row[2] is not None:
      oficina = row[2]
    
    month_year = row[0]
    material = row[3]
    descripcion = row[4]
    nivel_2 = row[5]

    stock_almacen_lib = row[6] or 0
    dias_almacen_centro_lib = row[7] or 0
    dias_almacen_lib = row[8] or 0

    stock_almacen_no_lib = row[9] or 0
    dias_almacen_centro_no_lib = row[10] or 0
    dias_almacen_no_lib = row[11] or 0

    stock_oficina_lib = row[15] or 0
    dias_oficina_centro_lib = row[16] or 0
    dias_oficina_lib = row[17] or 0

    stock_oficina_no_lib = row[18] or 0
    dias_oficina_centro_no_lib = row[19] or 0
    dias_oficina_no_lib = row[20] or 0

    llave = f'{oficina.lower()}{material}'

    if oficina.lower() in dict_lead_time['optimista']['Venta Local'].keys():
      ws[f'A{i}'].value = month_year
      ws[f'B{i}'].value = sector
      ws[f'C{i}'].value = oficina
      ws[f'D{i}'].value = int(material)
      ws[f'E{i}'].value = descripcion
      ws[f'F{i}'].value = nivel_2
      ws[f'G{i}'].value = stock_oficina_lib or 0
      ws[f'H{i}'].value = dias_oficina_centro_lib or 0
      ws[f'I{i}'].value = dias_oficina_lib or 0
      ws[f'K{i}'].value = stock_oficina_no_lib or 0
      ws[f'L{i}'].value = dias_oficina_centro_no_lib or 0
      ws[f'M{i}'].value = dias_oficina_no_lib or 0
      ws[f'O{i}'].value = stock_almacen_lib or 0
      ws[f'P{i}'].value = dias_almacen_centro_lib or 0
      ws[f'Q{i}'].value = dias_almacen_lib or 0
      ws[f'T{i}'].value = stock_almacen_no_lib or 0
      ws[f'U{i}'].value = dias_almacen_centro_no_lib or 0
      ws[f'V{i}'].value = dias_almacen_no_lib or 0

      today = datetime.now()
      holidays_country = holidays.CL(years=today.year)

      if 'america' in oficina.lower():
        holidays_country = holidays.US(years=today.year)        # USA
      elif 'europa' in oficina.lower():
        holidays_country = holidays.IT(years=today.year)        # Italia
      elif 'mexico' in oficina.lower():
        holidays_country = holidays.MX(years=today.year)        # Mexico
      elif 'shanghai'in oficina.lower():  
        holidays_country = holidays.CN(years=today.year)        # China
      elif 'asia' in oficina.lower():
        holidays_country = holidays.KR(years=today.year)        # Corea del sur

       # Considerar de lunes a sábado sin los feriados
      business_days = 0
      sm = 12
      for day in obj.itermonthdates(today.year, sm):
        if day.month == sm:
          if day not in holidays_country and day.strftime('%A') != 'Sunday':
            business_days += 1
      
      leftover_days = business_days - today.day 
      lead_time = dict_lead_time['optimista']['Venta Local'][oficina.lower()]
    
      # PUERTO OFICINA
      # Stock liberado
      dias_oficina = dias_oficina_centro_lib + leftover_days
      if dias_oficina >= lead_time['Destino']:
        ws[f'J{i}'].value = stock_oficina_lib
        ws[f'J{i}'].font = Font(bold=True, color=green)
        ws[f'J{i}'].fill = PatternFill("solid", fgColor=lightGreen)
      else:
        ws[f'J{i}'].value = 0
        ws[f'J{i}'].font = Font(bold=True, color=darkRed)
        ws[f'J{i}'].fill = PatternFill("solid", fgColor=lightRed)
      
      # Stock no liberado
      dias_oficina = dias_oficina_centro_no_lib + leftover_days
      if dias_oficina >= lead_time['Destino']:
        ws[f'N{i}'].value = stock_oficina_no_lib
        ws[f'N{i}'].font = Font(bold=True, color=green)
        ws[f'N{i}'].fill = PatternFill("solid", fgColor=lightGreen)
      else:
        ws[f'N{i}'].value = 0
        ws[f'N{i}'].font = Font(bold=True, color=darkRed)
        ws[f'N{i}'].fill = PatternFill("solid", fgColor=lightRed)
      
      # ALMACEN OFICINA
      # Stock liberado
      dias_almacen = dias_almacen_lib + leftover_days
      if dias_almacen >= lead_time['Almacen']:
        ws[f'R{i}'].value = stock_almacen_lib
        ws[f'R{i}'].font = Font(bold=True, color=green)
        ws[f'R{i}'].fill = PatternFill("solid", fgColor=lightGreen)
      else:
        ws[f'R{i}'].value = 0
        ws[f'R{i}'].font = Font(bold=True, color=darkRed)
        ws[f'R{i}'].fill = PatternFill("solid", fgColor=lightRed)
      
      dias_almacen = dias_almacen_no_lib + leftover_days
      if dias_almacen >= lead_time['Almacen']:
        ws[f'V{i}'].value = stock_almacen_no_lib
        ws[f'V{i}'].font = Font(bold=True, color=green)
        ws[f'V{i}'].fill = PatternFill("solid", fgColor=lightGreen)
      else:
        ws[f'V{i}'].value = 0
        ws[f'V{i}'].font = Font(bold=True, color=darkRed)
        ws[f'V{i}'].fill = PatternFill("solid", fgColor=lightRed)
      
      # STYLES
      line_blue = Side(border_style="thin", color=blue)
      line_grey = Side(border_style="thin", color=grey)
      thin = Side(border_style="thin", color=white)

      ws[f'G{i}'].border = Border(left=line_blue)
      ws[f'G{i}'].number_format = BUILTIN_FORMATS[3]
      ws[f'H{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'I{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'J{i}'].number_format = BUILTIN_FORMATS[3]
      
      ws[f'K{i}'].number_format = BUILTIN_FORMATS[3]
      ws[f'L{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'M{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'N{i}'].number_format = BUILTIN_FORMATS[3]

      ws[f'O{i}'].border = Border(left=line_blue)
      ws[f'O{i}'].number_format = BUILTIN_FORMATS[3]
      ws[f'P{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'Q{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'R{i}'].number_format = BUILTIN_FORMATS[3]

      ws[f'S{i}'].number_format = BUILTIN_FORMATS[3]
      ws[f'T{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'U{i}'].number_format = BUILTIN_FORMATS[2]
      ws[f'V{i}'].number_format = BUILTIN_FORMATS[3]
      ws[f'W{i}'].border = Border(left=line_blue)

      ws[f'A{i}'].font = Font(bold=False, color=blue)
      ws[f'B{i}'].font = Font(bold=False, color=blue)
      ws[f'C{i}'].font = Font(bold=False, color=blue)
      ws[f'D{i}'].font = Font(bold=False, color=blue)
      ws[f'E{i}'].font = Font(bold=False, color=blue)
      ws[f'F{i}'].font = Font(bold=False, color=blue)
      ws[f'A{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
      ws[f'B{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
      ws[f'C{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
      ws[f'D{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
      ws[f'E{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
      ws[f'F{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
      ws[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
      ws[f'B{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
      ws[f'C{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
      ws[f'D{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
      ws[f'E{i}'].border = Border(top=thin, left=thin, bottom=thin)
      i += 1  


  wb_dias_stock.close()    

  # ----- Cerramos y guardamos
  run_styles(ws)
  # run_number_format(ws)