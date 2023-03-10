from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, BUILTIN_FORMATS
from constants import *

def run_styles(ws):
  thin = Side(border_style="thin", color=white)
  ws['G1'].font = Font(bold=True, color=white)
  ws['G1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws['G1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws['G1'].fill = PatternFill("solid", fgColor=lightBlue)

  ws['O1'].font = Font(bold=True, color=white)
  ws['O1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws['O1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws['O1'].fill = PatternFill("solid", fgColor=lightBlue)

  for letter in ['G', 'K', 'O', 'S']:
    ws[f'{letter}2'].font = Font(bold=True, color=white)
    ws[f'{letter}2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'{letter}2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{letter}2'].fill = PatternFill("solid", fgColor=lightBlue)

  for i in range(1, ws.max_column + 1):
    ws[f'{get_column_letter(i)}3'].font = Font(bold=True, color=white)
    ws[f'{get_column_letter(i)}3'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'{get_column_letter(i)}3'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{get_column_letter(i)}3'].fill = PatternFill("solid", fgColor=lightBlue)

    ws.column_dimensions[f'{get_column_letter(i)}'].width = 10

  # Primera Fila
  ws.merge_cells('G1:N1')
  ws.merge_cells('O1:V1')

  # Segunda Fila
  ws.merge_cells('G2:J2')
  ws.merge_cells('K2:N2')
  ws.merge_cells('O2:R2')
  ws.merge_cells('S2:V2')

  # Tamaños
  ws.column_dimensions['C'].width = 16
  ws.column_dimensions['D'].width = 10
  ws.column_dimensions['E'].width = 33
  ws.column_dimensions['F'].width = 16

  # Tamaño
  ws.row_dimensions[1].height =  25
  ws.row_dimensions[2].height =  25


# def run_number_format(ws):
#   line_blue = Side(border_style="thin", color=blue)
#   line_grey = Side(border_style="thin", color=grey)
#   thin = Side(border_style="thin", color=white)

#   for i in range(4, ws.max_row + 1):
    # ws[f'F{i}'].border = Border(left=line_blue)
    # ws[f'F{i}'].number_format = BUILTIN_FORMATS[3]
    # ws[f'G{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'H{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'I{i}'].number_format = BUILTIN_FORMATS[3]
    
    # ws[f'J{i}'].number_format = BUILTIN_FORMATS[3]
    # ws[f'K{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'L{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'M{i}'].number_format = BUILTIN_FORMATS[3]

    # ws[f'N{i}'].border = Border(left=line_blue)
    # ws[f'N{i}'].number_format = BUILTIN_FORMATS[3]
    # ws[f'O{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'P{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'Q{i}'].number_format = BUILTIN_FORMATS[3]

    # ws[f'R{i}'].number_format = BUILTIN_FORMATS[3]
    # ws[f'S{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'T{i}'].number_format = BUILTIN_FORMATS[2]
    # ws[f'U{i}'].number_format = BUILTIN_FORMATS[3]
    # ws[f'V{i}'].border = Border(left=line_blue)

    # ws[f'A{i}'].font = Font(bold=False, color=blue)
    # ws[f'B{i}'].font = Font(bold=False, color=blue)
    # ws[f'C{i}'].font = Font(bold=False, color=blue)
    # ws[f'D{i}'].font = Font(bold=False, color=blue)
    # ws[f'E{i}'].font = Font(bold=False, color=blue)
    # ws[f'A{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'B{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'C{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'D{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'E{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # ws[f'B{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # ws[f'C{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # ws[f'D{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # ws[f'E{i}'].border = Border(top=thin, left=thin, bottom=thin)