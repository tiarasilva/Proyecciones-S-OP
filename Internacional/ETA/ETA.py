from openpyxl import load_workbook
from ETA.styles import *
from constants import *

import time
import calendar
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import holidays
start_time = time.time()

def create_ETA(ws, dict_lead_time, date_selected_month, dict_cierre_venta, dict_holidays, filename_logistica, filename_pedidos_confirmados):
  # DATA
  ws.append({
    # 6: f'=SUBTOTAL(9;Tabla14[Mes {selected_month}])'
    8: 'OPTIMISTA',
    18: 'PESIMISTA',
  })

  month_1 = date_selected_month
  month_2 = date_selected_month + relativedelta(months=1)
  month_3 = date_selected_month + relativedelta(months=2)
  month_4 = date_selected_month + relativedelta(months=3)

  name_month_1 = month_translate_EN_CL[month_1.strftime('%B').lower()]
  name_month_2 = month_translate_EN_CL[month_2.strftime('%B').lower()]
  name_month_3 = month_translate_EN_CL[month_3.strftime('%B').lower()]
  name_month_4 = month_translate_EN_CL[month_4.strftime('%B').lower()]

  selected_month1_year = f'{month_1.strftime("%m")}.{month_1.year}'
  selected_month2_year = f'{month_2.strftime("%m")}.{month_2.year}'
  selected_month3_year = f'{month_3.strftime("%m")}.{month_3.year}'
  selected_month4_year = f'{month_4.strftime("%m")}.{month_4.year}'


  ws.append({
    1: 'Sector',                          # A
    2: 'Canal de Distribución',           # B
    3: 'Pedido',                          # C
    4: 'Oficina',                         # D
    5: 'Material',                        # E
    6: 'Llave',                           # F
    7: 'ETA',                             # G
    8: f'Centro Agua {name_month_1}',     # H
    9: f'Centro Agua {name_month_2}',     # I
    10: f'Centro Agua {name_month_3}',    # J
    11: f'Centro Agua {name_month_4}',    # K
    12: 'Lead time 1',                    # L
    13: 'Fecha final 1',                  # M 
    14: 'Lead time 2',                    # N
    15: 'Fecha final 2',                  # O
    16: 'Leftover days',                  # P
    17: 'Considerar PES.',                # Q
    18: f'Centro Agua {name_month_1}',    # R
    19: f'Centro Agua {name_month_2}',    # S
    20: f'Centro Agua {name_month_3}',    # T
    21: f'Centro Agua {name_month_4}',    # U
    22: 'Lead time 1',                    # V
    23: 'Fecha final 1',                  # W
    24: 'Lead time 2',                    # X
    25: 'Fecha final 2',                  # Y
    26: 'Leftover days',                  # Z
    27: 'Considerar OPT',                 # AA
  })

  # ----- 3. File Confirmados
  wb_confirmados = load_workbook(filename_pedidos_confirmados, read_only=True, data_only=True)
  sheet_name = "CONF - AP (zarpe mes n y n+1) "
  dict_confirmados = {}

  if sheet_name in wb_confirmados.sheetnames:
    ws_confirmados = wb_confirmados[sheet_name]
  else:
    ws_confirmados = wb_confirmados.active
    print(f'En archivo {filename_pedidos_confirmados} no se encontró la hoja: "{sheet_name}", se utilizará la llamada: {wb_confirmados.active}')
  
  ws_confirmados_max = ws_confirmados.max_row

  i = 3
  print("--- %s ETA 5.3  ---" % (time.time() - start_time))
  for row in ws_confirmados.iter_rows(3, ws_confirmados_max, values_only=True):
    if row[0] is None:
      break
    oficina = row[0]
    meanwhile_oficina = oficina.split(' ', 1)
    codigo_oficina = meanwhile_oficina[0]
    oficina = meanwhile_oficina[1]

    sector = row[1]
    pedido = row[2]
    status = row[3]
    pto_destino = row[4]
    tipo_venta = row[5]
    material = row[6]
    descripcion = row[7]
    nivel_2 = row[8]
    fecha_programa = row[10]
    kilos_stock = row[12]
    kilos_proyectado = row[13]
    total_kilos = row[14]
    ETD = row[15]
    ETA = row[16]
    NAVE = row[17]
    fecha_carga = row[17]

    # !!!!! QUE HAGO CON LOS REPETIDOS!!! ---> Se borra el AP Confirmados
    key = f"{oficina.lower()}{pedido}{material}"

    if key in dict_confirmados:
      print(f"Repetido Confirmados {oficina}, pedido: {pedido}, material: {material}\n")
    
    else:
      dict_confirmados[key] = {
        'oficina': oficina,
        'tipo_venta': tipo_venta,
        'sector': sector,
        'pedido': pedido,
        'material': material,
        'descripcion': descripcion,
        'nivel 2': nivel_2,
        'kilos stock': kilos_stock,
        'kilos proyectado': kilos_proyectado,
        'total kilos': total_kilos,
        'ETD': ETD,
        'ETA': ETA,
      } 
    
    canal_distribucion = 'Venta Directa'

    if oficina.lower() in dict_lead_time['optimista']['Venta Local']:
      canal_distribucion = 'Venta Local'

    if ETA:
      ws[f'A{i}'].value = dict_sector_numero[int(sector)]
      ws[f'B{i}'].value = canal_distribucion
      ws[f'C{i}'].value = pedido
      ws[f'D{i}'].value = oficina
      ws[f'E{i}'].value = material
      ws[f'F{i}'].value = f'{oficina.lower()}{material}'
      ws[f'G{i}'].value = datetime.date(ETA)
      ws[f'H{i}'].value = total_kilos
      i += 1
    wb_confirmados.close()
  
  # ----- 4. File Logistica
  wb_logistica = load_workbook(filename_logistica, read_only=True, data_only=True)
  ws_logistica = wb_logistica['Pedidos Planta-Puerto-Embarcado']
  ws_logistica_max = ws_logistica.max_row
  dict_logistica = {}
  j = ws.max_row + 1

  print("--- %s ETA 5.4  ---" % (time.time() - start_time))
  for row in ws_logistica.iter_rows(2, ws_logistica_max, values_only=True):
    if row[1] is None:
      break
    oficina = row[1]
    tipo_venta = row[2]
    pedido = row[3]
    status_pedido = row[4]
    material = row[5]
    nave = row[6]
    pto_destino = row[7]
    fecha_despacho_real = row[8]
    ETD = row[9]
    ETA = row[10]
    naviera = row[11]
    kilos = row[12]
    ubicacion = row[13]
    key = f"{oficina.lower()}{pedido}{material}"
    canal_distribucion = 'Venta Directa'

    if oficina.lower() in dict_lead_time['optimista']['Venta Local']:
      canal_distribucion = 'Venta Local'

    if key in dict_confirmados.keys():
      # !!!!! QUE HAGO CON LOS REPETIDOS!!! ---> Se borra el AP Confirmados
      print(f"Repetido Confirmados con Logistica {oficina}, pedido: {pedido}, material: {material}")
      dict_confirmados.pop(key)

    else:
      if key in dict_logistica.keys():
        if ETA != dict_logistica[key]['ETA']:
          print(f"Repetido Logistica {oficina}, pedido: {pedido}, material: {material}")
        else:
          dict_logistica[key]['kilos'] += kilos
          kilos = dict_logistica[key]['kilos']

      if ETA and type(ETA) != str:
        # ws[f'A{j}'].value = dict_sector_numero[int(sector)]
        ws[f'B{j}'].value = canal_distribucion
        ws[f'C{j}'].value = pedido
        ws[f'D{j}'].value = oficina
        ws[f'E{j}'].value = material
        ws[f'F{j}'].value = f'{oficina.lower()}{material}'
        ws[f'G{j}'].value = datetime.date(ETA)
        ws[f'H{j}'].value = kilos
        j += 1
  
        dict_logistica[key] = {
          'oficina': oficina,
          'tipo_venta': tipo_venta,
          'sector': '',
          'pedido': pedido,
          'status_pedido': status_pedido,
          'material': material,
          'pto_destino': pto_destino,
          'fecha_despacho_real': fecha_despacho_real,
          'ETD': ETD,
          'ETA': ETA,
          'kilos': kilos,
          'ubicacion': ubicacion,
        }
  wb_logistica.close()

  # ----- 5. Calculo FECHAS
  dict_leftover_date = {}
  # Quiero todas las llaves que llegan en N, N+1 y N+2
  dict_ETA_sin_venta = {}

  print("--- %s ETA 5.5  ---" % (time.time() - start_time))
  for i, row in enumerate(ws.iter_rows(3, ws.max_row - 1, values_only=True), 3):
    sector = row[0]
    canal_distribucion = row[1]
    oficina = row[3]
    material = row[4]
    llave = row[5]
    ETA = row[6]
    kilos = ws[f'H{i}'].value
    ws[f'H{i}'].value = ''
    key_month_year = f'{ETA.strftime("%m")}.{ETA.year}'

    if canal_distribucion == "Venta Directa":
      if (ETA.month, ETA.year) == (month_1.month, month_1.year):
        ws[f'H{i}'].value = kilos
        ws[f'R{i}'].value = kilos
      elif (ETA.month, ETA.year)  == (month_2.month, month_2.year):
        ws[f'I{i}'].value = kilos
        ws[f'S{i}'].value = kilos
      elif (ETA.month, ETA.year)  == (month_3.month, month_3.year):
        ws[f'J{i}'].value = kilos
        ws[f'T{i}'].value = kilos
      elif (ETA.month, ETA.year) == (month_3.month, month_3.year):
        ws[f'K{i}'].value = kilos
        ws[f'O{i}'].fill = PatternFill("solid", fgColor=yellow)
        ws[f'U{i}'].value = kilos
        ws[f'U{i}'].fill = PatternFill("solid", fgColor=yellow)
    
    # CHECK ALL THE KEYS THAT ARE ON THE 3 MONTHS SELECTED PERIOD
    start_month_1 = date(month_1.year, month_1.month, 1)
    start_month_4 = date(month_4.year, month_4.month, 1)
    # key_month_year = f'{tiempo_final_pes.strftime("%m")}.{tiempo_final_pes.year}'
    # if key_month_year in dict_ETA_sin_venta.keys():
    if llave not in dict_ETA_sin_venta.keys() and start_month_1 <= ETA < start_month_4:
      dict_ETA_sin_venta[llave] = {
        'canal_distribucion': canal_distribucion,
        'oficina': oficina,
        'material': material,
        'sector': sector,
      }
        

    # ----- OPTIMISTA -----
    if canal_distribucion == "Venta Local":
      lead_time_opt = dict_lead_time['optimista'][canal_distribucion][oficina.lower()]
      LT_3 = lead_time_opt['Agua']
      LT_4 = lead_time_opt['Destino']
      LT_5 = lead_time_opt['Almacen']
      tiempo_final_opt = ETA + timedelta(LT_4) + timedelta(LT_5)
      ws[f'L{i}'].value = LT_4
      ws[f'M{i}'].value = ETA + timedelta(LT_4)
      ws[f'N{i}'].value = LT_5
      ws[f'O{i}'].value = tiempo_final_opt
        
      if tiempo_final_opt.month == month_1.month:
        ws[f'H{i}'].value = kilos
      
      elif tiempo_final_opt.month == month_2.month:
        ws[f'I{i}'].value = kilos
      
      elif tiempo_final_opt.month == month_3.month:
        ws[f'J{i}'].value = kilos
      
      elif tiempo_final_opt.month == month_4.month:
        ws[f'K{i}'].value = kilos
        ws[f'O{i}'].fill = PatternFill("solid", fgColor=yellow)
    
      # ----- PESIMISTA -----
      lead_time_pes = dict_lead_time['pesimista'][canal_distribucion][oficina.lower()]
      tiempo_final_pes = ETA + timedelta(LT_4) + timedelta(LT_5)
      LT_3 = lead_time_pes['Agua']
      LT_4 = lead_time_pes['Destino']
      LT_5 = lead_time_pes['Almacen']
      ws[f'V{i}'].value = LT_4
      ws[f'W{i}'].value = ETA + timedelta(LT_4)
      ws[f'X{i}'].value = LT_5
      ws[f'Y{i}'].value = tiempo_final_pes

      if tiempo_final_pes.month == month_1.month:
        ws[f'R{i}'].value = kilos
      
      elif tiempo_final_pes.month == month_2.month:
        ws[f'S{i}'].value = kilos
      
      elif tiempo_final_pes.month == month_3.month:
        ws[f'T{i}'].value = kilos
      
      elif tiempo_final_pes.month == month_4.month:
        ws[f'U{i}'].value = kilos
        ws[f'U{i}'].fill = PatternFill("solid", fgColor=yellow)
      
      # CHECK ALL THE KEYS THAT ARE ON THE 3 MONTHS SELECTED PERIOD
      start_month_1 = date(month_1.year, month_1.month, 1)
      start_month_4 = date(month_4.year, month_4.month, 1)
      # key_month_year = f'{tiempo_final_pes.strftime("%m")}.{tiempo_final_pes.year}'
      # if key_month_year in dict_ETA_sin_venta.keys():
      if llave not in dict_ETA_sin_venta.keys() and start_month_1 <= tiempo_final_pes < start_month_4:
        dict_ETA_sin_venta[llave] = {
          'canal_distribucion': canal_distribucion,
          'oficina': oficina,
          'material': material,
          'sector': sector,
        }

    # ---- Calendario leftover days for month
      for tiempo_final in [tiempo_final_pes, tiempo_final_opt]:
        holidays_country = dict_holidays[tiempo_final.year]['chile']
        if canal_distribucion == 'Venta Local':
          holidays_country = dict_holidays[tiempo_final.year][oficina.lower()]
        leftover_days = 0

        if oficina.lower() in dict_leftover_date.keys():
          if tiempo_final in dict_leftover_date[oficina.lower()]:
            leftover_days = dict_leftover_date[oficina.lower()][tiempo_final]

          else:
            last_day_month = calendar.monthrange(tiempo_final.year, tiempo_final.month)[1]
            for day in range(tiempo_final.day + 1, last_day_month + 1):
              date_day = date(tiempo_final.year, tiempo_final.month, day)
              if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
                leftover_days += 1
            dict_leftover_date[oficina.lower()][tiempo_final] = leftover_days

        else:
          dict_leftover_date[oficina.lower()] = {}
          last_day_month = calendar.monthrange(tiempo_final.year, tiempo_final.month)[1]
          for day in range(tiempo_final.day + 1, last_day_month + 1):
            date_day = date(tiempo_final.year, tiempo_final.month, day)
            if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
              leftover_days += 1
          dict_leftover_date[oficina.lower()][tiempo_final] = leftover_days

      leftover_opt = dict_leftover_date[oficina.lower()][tiempo_final_opt]
      ws[f'P{i}'].value = leftover_opt

      if leftover_opt > dict_cierre_venta[oficina.lower()]:
        ws[f'Q{i}'].value = 'SI'
      else:
        ws[f'Q{i}'].value = f'Mes {tiempo_final_opt.month + 1}'
      
      leftover_pes = dict_leftover_date[oficina.lower()][tiempo_final_pes]
      ws[f'Z{i}'].value = leftover_pes

      if leftover_pes > dict_cierre_venta[oficina.lower()]:
        ws[f'AA{i}'].value = 'SI'
      else:
        ws[f'AA{i}'].value = f'Mes {tiempo_final_pes.month + 1}'
    
    # NUMBER FORMAT
    ws[f'H{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'I{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'J{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'K{i}'].number_format = BUILTIN_FORMATS[3]

    ws[f'L{i}'].number_format = BUILTIN_FORMATS[4]
    ws[f'N{i}'].number_format = BUILTIN_FORMATS[4]

    ws[f'R{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'S{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'T{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'U{i}'].number_format = BUILTIN_FORMATS[3]

    ws[f'V{i}'].number_format = BUILTIN_FORMATS[4]
    ws[f'X{i}'].number_format = BUILTIN_FORMATS[4]

    ws[f'Q{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'AA{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # # STYLES
    # thin = Side(border_style="thin", color=white)
    # line_blue = Side(border_style="thin", color=blue)

    # ws[f'A{i}'].font = Font(bold=False, color=blue)
    # ws[f'A{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'B{i}'].font = Font(bold=False, color=blue)
    # ws[f'B{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'B{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'C{i}'].font = Font(bold=False, color=blue)
    # ws[f'C{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'C{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'D{i}'].font = Font(bold=False, color=blue)
    # ws[f'D{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'D{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'E{i}'].font = Font(bold=False, color=blue)
    # ws[f'E{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'E{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'F{i}'].font = Font(bold=False, color=blue)
    # ws[f'F{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'F{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'G{i}'].font = Font(bold=False, color=blue)
    # ws[f'G{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'G{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  
  # dict_ETAS = dict_confirmados | dict_logistica
  # return dict_ETAS
  
  print("--- %s ETA 5 ---" % (time.time() - start_time))
# ----- Corremos estilos y cerramos
  run_styles(ws)
  # run_number_format(ws)
  print("--- %s ETA 6 ---" % (time.time() - start_time))

  return dict_ETA_sin_venta