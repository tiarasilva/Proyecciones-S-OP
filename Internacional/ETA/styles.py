from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from constants import *

import time
start_time = time.time()

def run_styles(ws):
  thin = Side(border_style="thin", color=white)

  for letter in range(1, 8):
    letter = get_column_letter(letter)
    a_col = ws.column_dimensions[letter]
    a_col.font = Font(bold=False, color=blue)
    a_col.fill = PatternFill("solid", fgColor=lightlightBlue)
    a_col.border = Border(top=thin, left=thin, right=thin, bottom=thin)

  for i in range(1, ws.max_column + 1):
    ws[f'{get_column_letter(i)}2'].font = Font(bold=True, color=white)
    ws[f'{get_column_letter(i)}2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'{get_column_letter(i)}2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{get_column_letter(i)}2'].fill = PatternFill("solid", fgColor=lightBlue)

  # Tamaños
  ws.column_dimensions['A'].width = 10
  ws.column_dimensions['B'].width = 16
  ws.column_dimensions['C'].width = 10
  ws.column_dimensions['D'].width = 19
  ws.column_dimensions['E'].width = 10
  ws.column_dimensions['F'].width = 22
  ws.column_dimensions['G'].width = 12
  ws.column_dimensions['M'].width = 12

  ws.column_dimensions['O'].width = 12
  ws.column_dimensions['Q'].width = 12

  ws.column_dimensions['W'].width = 12
  ws.column_dimensions['Y'].width = 12
  ws.column_dimensions['AA'].width = 12

  ws.row_dimensions[1].height = 25

  # Merge
  ws.merge_cells('H1:Q1')
  ws.merge_cells('R1:AA1')

  ws['H1'].font = Font(bold=True, color=white)
  ws['H1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws['H1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws['H1'].fill = PatternFill("solid", fgColor=lightBlue)
  
  ws['R1'].font = Font(bold=True, color=white)
  ws['R1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws['R1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws['R1'].fill = PatternFill("solid", fgColor=lightBlue)

def run_number_format(ws):
  max_row = ws.max_row
  thin = Side(border_style="thin", color=white)
  line_blue = Side(border_style="thin", color=blue)

  print("--- %s ETA 6.3 ---" % (time.time() - start_time))

  for i in range(3, max_row + 1):
    ws[f'H{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'I{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'J{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'K{i}'].number_format = BUILTIN_FORMATS[3]

    ws[f'L{i}'].number_format = BUILTIN_FORMATS[4]
    ws[f'N{i}'].number_format = BUILTIN_FORMATS[4]

    ws[f'R{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'S{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'T{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'U{i}'].number_format = BUILTIN_FORMATS[3]

    ws[f'V{i}'].number_format = BUILTIN_FORMATS[4]
    ws[f'X{i}'].number_format = BUILTIN_FORMATS[4]

    # ws[f'A{i}'].font = Font(bold=False, color=blue)
    # ws[f'A{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'B{i}'].font = Font(bold=False, color=blue)
    # ws[f'B{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'B{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'C{i}'].font = Font(bold=False, color=blue)
    # ws[f'C{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'C{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'D{i}'].font = Font(bold=False, color=blue)
    # ws[f'D{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'D{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'E{i}'].font = Font(bold=False, color=blue)
    # ws[f'E{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'E{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws[f'F{i}'].font = Font(bold=False, color=blue)
    # ws[f'F{i}'].fill = PatternFill("solid", fgColor=lightlightBlue)
    # ws[f'F{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws[f'H{i}'].border = Border(left=line_blue)
    ws[f'L{i}'].border = Border(left=line_blue)
    ws[f'R{i}'].border = Border(left=line_blue)
    ws[f'V{i}'].border = Border(left=line_blue)
    ws[f'AB{i}'].border = Border(left=line_blue)

    ws[f'L{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'Q{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'P{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'V{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'Z{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f'AA{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  print("--- %s ETA 6.4 ---" % (time.time() - start_time))