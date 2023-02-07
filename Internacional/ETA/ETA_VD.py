from openpyxl import load_workbook
from ETA.styles import *
from constants import *

import time
import calendar
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import holidays
start_time = time.time()

def create_ETA_VD(ws, dict_lead_time, date_selected_month, dict_cierre_venta, dict_holidays):
  # HOLIDAYS
  today = datetime.now()
  last_year = today.year - 1
  this_year = today.year
  next_year = today.year + 1

  month_1 = date_selected_month
  month_2 = date_selected_month + relativedelta(months=1)
  month_3 = date_selected_month + relativedelta(months=2)
  month_4 = date_selected_month + relativedelta(months=3)

  name_month_1 = month_translate_EN_CL[month_1.strftime('%B').lower()]
  name_month_2 = month_translate_EN_CL[month_2.strftime('%B').lower()]
  name_month_3 = month_translate_EN_CL[month_3.strftime('%B').lower()]
  name_month_4 = month_translate_EN_CL[month_4.strftime('%B').lower()]

  ws.append({
    1: 'Sector',                            # A
    2: 'Pedido',                            # B
    3: 'Oficina',                           # C
    4: 'Material',                          # D
    5: 'Llave',                             # E
    6: 'ETA',                               # F
    7: f'Centro Agua {name_month_1}',       # G
    8: f'Centro Agua {name_month_2}',       # H
    9: f'Centro Agua {name_month_3}',       # I
    10: f'Centro Agua {name_month_4}',      # J
    11: 'Leftover days',                    # K
    12: 'Considerar PES.',                  # L
  })

  # ----- Filtrar información
  wb = load_workbook(filename_ETA, read_only=True, data_only=True)
  ws_ETA = wb['ETA']
  ws_ETA_max_row = ws_ETA.max_row
  dict_status = {
    'despachado': 'Puerto',
    'embarcado': 'Agua',
    'a programar': 'Planta',
    'programado': 'Planta'
  }
  dict_leftover_date_opt = {}
  dict_leftover_date_pes = {}
  dict_leftover_date = {}

  i = 2
  print("--- %s ETA 4.1 ---" % (time.time() - start_time))
  for row in ws_ETA.iter_rows(4, ws_ETA_max_row, values_only=True):
    pedido = row[0]
    eta = row[10]
    material = row[11]
    oficina = row[21]
    n_sector = row[28]
    kilos = row[33]
    eta = datetime.date(eta)

    if oficina.lower() in dict_lead_time['optimista']['directa']:
      ws[f'A{i}'].value = dict_sector_numero[n_sector]
      ws[f'B{i}'].value = pedido
      ws[f'C{i}'].value = oficina
      ws[f'D{i}'].value = material
      ws[f'E{i}'].value = f'{oficina.lower()}{material}'
      ws[f'F{i}'].value = eta


      if eta.month == month_1.month:
        ws[f'G{i}'].value = kilos
      
      elif eta.month == month_2.month:
        ws[f'H{i}'].value = kilos
      
      elif eta.month == month_3.month:
        ws[f'I{i}'].value = kilos
      
      else:
        ws[f'J{i}'].value = kilos
        ws[f'F{i}'].fill = PatternFill("solid", fgColor=yellow)

      # ---- Calendario leftover days for month
      holidays_country = dict_holidays[eta.year]['chile']
      leftover_days = 0

      if oficina.lower() in dict_leftover_date.keys():
        if eta in dict_leftover_date[oficina.lower()]:
          leftover_days = dict_leftover_date[oficina.lower()][eta]
        else:
          last_day_month = calendar.monthrange(eta.year, eta.month)[1]
          for day in range(eta.day + 1, last_day_month + 1):
            date_day = date(eta.year, eta.month, day)
            if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
              leftover_days += 1
          dict_leftover_date[oficina.lower()][eta] = leftover_days
      else:
        dict_leftover_date[oficina.lower()] = {}
        last_day_month = calendar.monthrange(eta.year, eta.month)[1]
        for day in range(eta.day + 1, last_day_month + 1):
          date_day = date(eta.year, eta.month, day)
          if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
            leftover_days += 1
        dict_leftover_date[oficina.lower()][eta] = leftover_days

      leftover = dict_leftover_date[oficina.lower()][eta]
      ws[f'K{i}'].value = leftover

      if leftover > dict_cierre_venta[oficina.lower()]:
        ws[f'L{i}'].value = 'SI'
      else:
        ws[f'L{i}'].value = f'Mes {eta.month + 1}'

      i += 1
  
  print("--- %s ETA 5 ---" % (time.time() - start_time))
# ----- Corremos estilos y cerramos
  wb.close()
  run_styles(ws, 1)
  print("--- %s ETA 6.1 ---" % (time.time() - start_time))
  run_number_format(ws)
  print("--- %s ETA 6 ---" % (time.time() - start_time))