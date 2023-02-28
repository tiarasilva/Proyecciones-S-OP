from openpyxl import load_workbook
from ETA.styles import *
from constants import *

import time
import calendar
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import holidays
start_time = time.time()

def create_ETA(ws, dict_lead_time, date_selected_month, dict_cierre_venta, dict_holidays):
  # DATA
  ws.append({
    # 6: f'=SUBTOTALES(9;Tabla14[Mes {selected_month}])'
    8: 'OPTIMISTA',
    18: 'PESIMISTA',
  })
  month_1 = date_selected_month
  month_2 = date_selected_month + relativedelta(months=1)
  month_3 = date_selected_month + relativedelta(months=2)
  month_4 = date_selected_month + relativedelta(months=3)

  name_month_1 = month_translate_EN_CL[month_1.strftime('%B').lower()]
  name_month_2 = month_translate_EN_CL[month_2.strftime('%B').lower()]
  name_month_3 = month_translate_EN_CL[month_3.strftime('%B').lower()]
  name_month_4 = month_translate_EN_CL[month_4.strftime('%B').lower()]

  ws.append({
    1: 'Sector',                          # A
    2: 'Canal de Distribución',           # B
    3: 'Pedido',                          # C
    4: 'Oficina',                         # D
    5: 'Material',                        # E
    6: 'Llave',                           # F
    7: 'ETA',                             # G
    8: f'Centro Agua {name_month_1}',     # H
    9: f'Centro Agua {name_month_2}',     # I
    10: f'Centro Agua {name_month_3}',    # J
    11: f'Centro Agua {name_month_4}',    # K
    12: 'Lead time 1',              # L
    13: 'Fecha final 1',            # M 
    14: 'Lead time 2',              # N
    15: 'Fecha final 2',            # O
    16: 'Leftover days',                  # P
    17: 'Considerar PES.',                # Q
    18: f'Centro Agua {name_month_1}',    # R
    19: f'Centro Agua {name_month_2}',    # S
    20: f'Centro Agua {name_month_3}',    # T
    21: f'Centro Agua {name_month_4}',    # U
    22: 'Lead time 1',              # V
    23: 'Fecha final 1',            # W
    24: 'Lead time 2',              # X
    25: 'Fecha final 2',            # Y
    26: 'Leftover days',                  # Z
    27: 'Considerar OPT',                 # AA
  })

  run_styles(ws)

  # ----- Filtrar información
  wb = load_workbook(filename_ETA, read_only=True, data_only=True)
  ws_ETA = wb['ETA']
  ws_ETA_max_row = ws_ETA.max_row
  dict_leftover_date = {}

  i = 3
  print("--- %s ETA 4.1 ---" % (time.time() - start_time))
  for row in ws_ETA.iter_rows(2, ws_ETA_max_row, values_only=True):
    pedido = row[0]
    eta = row[10]
    material = row[11]
    oficina = row[20]
    n_sector = row[27]
    kilos = row[33]
    eta = datetime.date(eta)
    canal_distribucion = 'Venta Directa'

    if oficina.lower() in dict_lead_time['optimista']['Venta Local']:
      canal_distribucion = 'Venta Local'
      
    ws[f'A{i}'].value = dict_sector_numero[n_sector]
    ws[f'B{i}'].value = canal_distribucion
    ws[f'C{i}'].value = pedido
    ws[f'D{i}'].value = oficina
    ws[f'E{i}'].value = material
    ws[f'F{i}'].value = f'{oficina.lower()}{material}'
    ws[f'G{i}'].value = eta

    # OPTIMISTA
    lead_time_opt = dict_lead_time['optimista'][canal_distribucion][oficina.lower()]
    LT_1 = lead_time_opt['Planta']
    LT_2 = lead_time_opt['Puerto']
    tiempo_final_opt = eta + timedelta(LT_1) + timedelta(LT_2)
    ws[f'L{i}'].value = LT_1
    ws[f'M{i}'].value = eta + timedelta(LT_1)
    ws[f'N{i}'].value = LT_2

    if canal_distribucion == 'Venta Local':
      LT_3 = lead_time_opt['Agua']
      LT_4 = lead_time_opt['Destino']
      LT_5 = lead_time_opt['Almacen']
      tiempo_final_opt = eta + timedelta(LT_4) + timedelta(LT_5)
      ws[f'L{i}'].value = LT_4
      ws[f'M{i}'].value = eta + timedelta(LT_4)
      ws[f'N{i}'].value = LT_5
      
    if tiempo_final_opt.month == month_1.month:
      ws[f'H{i}'].value = kilos
    
    elif tiempo_final_opt.month == month_2.month:
      ws[f'I{i}'].value = kilos
    
    elif tiempo_final_opt.month == month_3.month:
      ws[f'J{i}'].value = kilos
    
    else:
      ws[f'K{i}'].value = kilos
      ws[f'O{i}'].fill = PatternFill("solid", fgColor=yellow)

    ws[f'O{i}'].value = tiempo_final_opt
    
    # PESIMISTA
    lead_time_pes = dict_lead_time['pesimista'][canal_distribucion][oficina.lower()]
    LT_1 = lead_time_pes['Planta']
    LT_2 = lead_time_pes['Puerto']
    tiempo_final_pes = eta + timedelta(LT_1) + timedelta(LT_2)
    ws[f'V{i}'].value = LT_1
    ws[f'W{i}'].value = eta + timedelta(LT_1)
    ws[f'X{i}'].value = LT_2

    if canal_distribucion == 'Venta Local':
      LT_3 = lead_time_pes['Agua']
      LT_4 = lead_time_pes['Destino']
      LT_5 = lead_time_pes['Almacen']
      ws[f'V{i}'].value = LT_4
      ws[f'W{i}'].value = eta + timedelta(LT_4)
      ws[f'X{i}'].value = LT_5

    if tiempo_final_pes.month == month_1.month:
      ws[f'R{i}'].value = kilos
    
    elif tiempo_final_pes.month == month_2.month:
      ws[f'S{i}'].value = kilos
    
    elif tiempo_final_pes.month == month_3.month:
      ws[f'T{i}'].value = kilos
    
    else:
      ws[f'U{i}'].value = kilos
      ws[f'U{i}'].fill = PatternFill("solid", fgColor=yellow)

    ws[f'Y{i}'].value = tiempo_final_pes

    # ---- Calendario leftover days for month
    for tiempo_final in [tiempo_final_pes, tiempo_final_opt]:
      holidays_country = dict_holidays[tiempo_final.year]['chile']
      if canal_distribucion == 'Venta Local':
        holidays_country = dict_holidays[tiempo_final.year][oficina.lower()]
      leftover_days = 0

      if oficina.lower() in dict_leftover_date.keys():
        if tiempo_final in dict_leftover_date[oficina.lower()]:
          leftover_days = dict_leftover_date[oficina.lower()][tiempo_final]
        else:
          last_day_month = calendar.monthrange(tiempo_final.year, tiempo_final.month)[1]
          for day in range(tiempo_final.day + 1, last_day_month + 1):
            date_day = date(tiempo_final.year, tiempo_final.month, day)
            if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
              leftover_days += 1
          dict_leftover_date[oficina.lower()][tiempo_final] = leftover_days
      else:
        dict_leftover_date[oficina.lower()] = {}
        last_day_month = calendar.monthrange(tiempo_final.year, tiempo_final.month)[1]
        for day in range(tiempo_final.day + 1, last_day_month + 1):
          date_day = date(tiempo_final.year, tiempo_final.month, day)
          if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
            leftover_days += 1
        dict_leftover_date[oficina.lower()][tiempo_final] = leftover_days

    leftover_opt = dict_leftover_date[oficina.lower()][tiempo_final_opt]

    ws[f'P{i}'].value = leftover_opt

    if leftover_opt > dict_cierre_venta[oficina.lower()]:
      ws[f'Q{i}'].value = 'SI'
    else:
      ws[f'Q{i}'].value = f'Mes {tiempo_final_opt.month + 1}'
    
    leftover_pes = dict_leftover_date[oficina.lower()][tiempo_final_pes]
    ws[f'Z{i}'].value = leftover_pes

    if leftover_pes > dict_cierre_venta[oficina.lower()]:
      ws[f'AA{i}'].value = 'SI'
    else:
      ws[f'AA{i}'].value = f'Mes {tiempo_final_pes.month + 1}'

    i += 1
  
  print("--- %s ETA 5 ---" % (time.time() - start_time))
# ----- Corremos estilos y cerramos
  wb.close()
  print("--- %s ETA 6.1 ---" % (time.time() - start_time))
  run_number_format(ws)
  print("--- %s ETA 6 ---" % (time.time() - start_time))