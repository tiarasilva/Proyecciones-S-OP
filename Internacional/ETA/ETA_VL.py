from openpyxl import load_workbook
from ETA.styles import *
from constants import *

import time
import calendar
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import holidays
start_time = time.time()

def create_ETA_VL(ws, dict_lead_time, date_selected_month, dict_cierre_venta, dict_holidays):

  # DATA
  ws.append({
    # 6: f'=SUBTOTALES(9;Tabla14[Mes {selected_month}])'
    7: 'OPTIMISTA',
    17: 'PESIMISTA',
  })

  month_1 = date_selected_month
  month_2 = date_selected_month + relativedelta(months=1)
  month_3 = date_selected_month + relativedelta(months=2)
  month_4 = date_selected_month + relativedelta(months=3)

  name_month_1 = month_translate_EN_CL[month_1.strftime('%B').lower()]
  name_month_2 = month_translate_EN_CL[month_2.strftime('%B').lower()]
  name_month_3 = month_translate_EN_CL[month_3.strftime('%B').lower()]
  name_month_4 = month_translate_EN_CL[month_4.strftime('%B').lower()]

  ws.append({
    1: 'Sector', 
    2: 'Pedido',
    3: 'Oficina', 
    4: 'Material', 
    5: 'Llave',
    6: 'ETA',
    7: f'Centro Agua {name_month_1}',
    8: f'Centro Agua {name_month_2}',
    9: f'Centro Agua {name_month_3}',
    10: f'Centro Agua {name_month_4}',
    11: 'Lead time Destino',
    12: 'Fecha final Destino',
    13: 'Lead time Almacen',
    14: 'Fecha final Almacen',
    15: 'Leftover days',
    16: 'Considerar PES.',
    17: f'Centro Agua {name_month_1}',
    18: f'Centro Agua {name_month_2}',
    19: f'Centro Agua {name_month_3}',
    20: f'Centro Agua {name_month_4}',
    21: 'Lead time Destino',
    22: 'Fecha final Destino',
    23: 'Lead time Almacen',
    24: 'Fecha final Almacen',
    25: 'Leftover days',
    26: 'Considerar OPT',
  })

  # ----- Filtrar información
  wb = load_workbook(filename_ETA, read_only=True, data_only=True)
  ws_ETA = wb['ETA']
  ws_ETA_max_row = ws_ETA.max_row
  dict_status = {
    'despachado': 'Puerto',
    'embarcado': 'Agua',
    'a programar': 'Planta',
    'programado': 'Planta'
  }
  dict_leftover_date_opt = {}
  dict_leftover_date_pes = {}
  dict_leftover_date = {}

  i = 3
  print("--- %s ETA 4.1 ---" % (time.time() - start_time))
  for row in ws_ETA.iter_rows(4, ws_ETA_max_row, values_only=True):
    pedido = row[0]
    eta = row[10]
    material = row[11]
    oficina = row[20]
    n_sector = row[27]
    kilos = row[33]
    eta = datetime.date(eta)

    if oficina.lower() in dict_lead_time['optimista']['local']:
      ws[f'A{i}'].value = dict_sector_numero[n_sector]
      ws[f'B{i}'].value = pedido
      ws[f'C{i}'].value = oficina
      ws[f'D{i}'].value = material
      ws[f'E{i}'].value = f'{oficina.lower()}{material}'
      ws[f'F{i}'].value = eta

      # OPTIMISTA
      lead_time_opt = dict_lead_time['optimista']['local'][oficina.lower()]
      LT_destino = lead_time_opt['Destino']
      LT_almacen = lead_time_opt['Almacen']
      tiempo_final_opt = eta + timedelta(LT_destino) + timedelta(LT_almacen)

      if tiempo_final_opt.month == month_1.month:
        ws[f'G{i}'].value = kilos
      
      elif tiempo_final_opt.month == month_2.month:
        ws[f'H{i}'].value = kilos
      
      elif tiempo_final_opt.month == month_3.month:
        ws[f'I{i}'].value = kilos
      
      else:
        ws[f'J{i}'].value = kilos
        ws[f'N{i}'].fill = PatternFill("solid", fgColor=yellow)

      ws[f'K{i}'].value = LT_destino
      ws[f'L{i}'].value = eta + timedelta(LT_destino)
      ws[f'M{i}'].value = LT_almacen
      ws[f'N{i}'].value = tiempo_final_opt
      
      # PESIMISTA
      lead_time_pes = dict_lead_time['pesimista']['local'][oficina.lower()]
      LT_destino = lead_time_pes['Destino']
      LT_almacen = lead_time_pes['Almacen']
      tiempo_final_pes = eta + timedelta(LT_destino) + timedelta(LT_almacen)

      if tiempo_final_pes.month == month_1.month:
        ws[f'Q{i}'].value = kilos
      
      elif tiempo_final_pes.month == month_2.month:
        ws[f'R{i}'].value = kilos
      
      elif tiempo_final_pes.month == month_3.month:
        ws[f'S{i}'].value = kilos
      
      else:
        ws[f'T{i}'].value = kilos
        ws[f'T{i}'].fill = PatternFill("solid", fgColor=yellow)

      ws[f'U{i}'].value = LT_destino
      ws[f'V{i}'].value = eta + timedelta(LT_destino)
      ws[f'W{i}'].value = LT_almacen
      ws[f'X{i}'].value = tiempo_final_pes

    # ---- Calendario leftover days for month
      for tiempo_final in [tiempo_final_pes, tiempo_final_opt]:
        holidays_country = dict_holidays[tiempo_final.year][oficina.lower()]
        leftover_days = 0

        if oficina.lower() in dict_leftover_date.keys():
          if tiempo_final in dict_leftover_date[oficina.lower()]:
            leftover_days = dict_leftover_date[oficina.lower()][tiempo_final]
          else:
            last_day_month = calendar.monthrange(tiempo_final.year, tiempo_final.month)[1]
            for day in range(tiempo_final.day + 1, last_day_month + 1):
              date_day = date(tiempo_final.year, tiempo_final.month, day)
              if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
                leftover_days += 1
            dict_leftover_date[oficina.lower()][tiempo_final] = leftover_days
        else:
          dict_leftover_date[oficina.lower()] = {}
          last_day_month = calendar.monthrange(tiempo_final.year, tiempo_final.month)[1]
          for day in range(tiempo_final.day + 1, last_day_month + 1):
            date_day = date(tiempo_final.year, tiempo_final.month, day)
            if date_day not in holidays_country and date_day.strftime('%A') != 'Sunday':
              leftover_days += 1
          dict_leftover_date[oficina.lower()][tiempo_final] = leftover_days

      leftover_opt = dict_leftover_date[oficina.lower()][tiempo_final_opt]
      ws[f'O{i}'].value = leftover_opt

      if leftover_opt > dict_cierre_venta[oficina.lower()]:
        ws[f'P{i}'].value = 'SI'
      else:
        ws[f'P{i}'].value = f'Mes {tiempo_final_opt.month + 1}'
      
      leftover_pes = dict_leftover_date[oficina.lower()][tiempo_final_pes]
      ws[f'Y{i}'].value = leftover_pes

      if leftover_pes > dict_cierre_venta[oficina.lower()]:
        ws[f'Z{i}'].value = 'SI'
      else:
        ws[f'Z{i}'].value = f'Mes {tiempo_final_pes.month + 1}'

      i += 1
  
  print("--- %s ETA 5 ---" % (time.time() - start_time))
# ----- Corremos estilos y cerramos
  wb.close()
  run_styles(ws, 2)
  VL_styles(ws)
  print("--- %s ETA 6.1 ---" % (time.time() - start_time))
  run_number_format(ws)
  print("--- %s ETA 6 ---" % (time.time() - start_time))